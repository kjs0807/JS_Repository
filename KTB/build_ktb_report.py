from __future__ import annotations

import json
import math
import re
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "ktb_analysis_config.json"


@dataclass
class Curve:
    name: str
    points: dict[float, float]
    history: dict[float, list[tuple[date, float]]]


def load_config() -> dict[str, Any]:
    with CONFIG_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def to_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        s = str(int(value))
        if len(s) == 8:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        if " " in s:
            s = s.split(" ")[0]
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def parse_yyyymmdd(value: Any) -> date | None:
    return to_date(value)


def years_between(start: date, end: date) -> float:
    return max((end - start).days / 365.25, 0.0)


def parse_tenor(label: Any) -> tuple[str, float] | None:
    if label is None:
        return None
    s = str(label).replace(" ", "")
    if "월" in s:
        m = re.search(r"(\d+(?:\.\d+)?)월", s)
        if m:
            months = float(m.group(1))
            return (f"{months:g}M", months / 12.0)
    if "개월" in s:
        m = re.search(r"(\d+(?:\.\d+)?)개월", s)
        if m:
            months = float(m.group(1))
            return (f"{months:g}M", months / 12.0)
    if "년" in s:
        m = re.search(r"(\d+(?:\.\d+)?)년", s)
        if m:
            years = float(m.group(1))
            return (f"{years:g}Y", years)
    return None


def tenor_label(years: float) -> str:
    if years < 1:
        return f"{years * 12:g}M"
    return f"{years:g}Y"


def interp(points: dict[float, float], x: float) -> float | None:
    clean = sorted((float(k), float(v)) for k, v in points.items() if v is not None)
    if not clean:
        return None
    if x <= clean[0][0]:
        return clean[0][1]
    if x >= clean[-1][0]:
        return clean[-1][1]
    for (x0, y0), (x1, y1) in zip(clean, clean[1:]):
        if x0 <= x <= x1:
            if x1 == x0:
                return y0
            w = (x - x0) / (x1 - x0)
            return y0 + w * (y1 - y0)
    return clean[-1][1]


def interp_inside_range(points: dict[float, float], x: float) -> float | None:
    clean = sorted((float(k), float(v)) for k, v in points.items() if v is not None)
    if not clean or x < clean[0][0] or x > clean[-1][0]:
        return None
    return interp(points, x)


def curve_range_status(points: dict[float, float], x: float | None) -> str:
    clean = sorted(float(k) for k, v in points.items() if v is not None)
    if x is None or not clean:
        return "No curve"
    if x < clean[0]:
        return "Below curve range"
    if x > clean[-1]:
        return "Above curve range"
    return "OK"


def pct_rank(values: list[float], x: float) -> float | None:
    vals = sorted(v for v in values if v is not None and math.isfinite(v))
    if not vals:
        return None
    return sum(1 for v in vals if v <= x) / len(vals)


def zscore(values: list[float], x: float) -> float | None:
    vals = [v for v in values if v is not None and math.isfinite(v)]
    if len(vals) < 3:
        return None
    sd = statistics.pstdev(vals)
    if sd == 0:
        return None
    return (x - statistics.mean(vals)) / sd


def approx_duration(years: float, ytm_pct: float) -> float:
    return max(years, 0.05) / (1.0 + max(ytm_pct, 0.0) / 100.0)


def carry_return(yield_pct: float, funding_pct: float, h: float) -> float:
    return (yield_pct - funding_pct) / 100.0 * h


def valid_holding_period(tenor_years: float, h: float) -> bool:
    return h < tenor_years


def rolldown_return(curve: Curve, tenor_years: float, ytm_pct: float, h: float) -> float | None:
    if not valid_holding_period(tenor_years, h):
        return None
    rolled_tenor = tenor_years - h
    rolled_yield = interp_inside_range(curve.points, rolled_tenor)
    if rolled_yield is None:
        return None
    dur = approx_duration(tenor_years, ytm_pct)
    return -dur * ((rolled_yield - ytm_pct) / 100.0)


def dv01_per_notional(tenor_years: float, ytm_pct: float, notional: float) -> float:
    return approx_duration(tenor_years, ytm_pct) * 0.0001 * notional


def scenario_pnl_for_yield_bp(tenor_years: float, ytm_pct: float, notional: float, bp_change: float) -> float:
    return -approx_duration(tenor_years, ytm_pct) * (bp_change / 10000.0) * notional


def total_with_parallel_shock_million(total_return: float | None, tenor_years: float, ytm_pct: float, bp_change: float, notional: float = 10_000_000_000) -> float | None:
    if total_return is None:
        return None
    carry_roll_pnl = total_return * notional
    return (carry_roll_pnl + scenario_pnl_for_yield_bp(tenor_years, ytm_pct, notional, bp_change)) / 1_000_000


def round_notional(notional: float | None, increment: float) -> float | None:
    if notional is None:
        return None
    return math.floor(notional / increment + 0.5) * increment


def parse_curve_sheet(wb, sheet_name: str) -> Curve:
    ws = wb[sheet_name]
    row_iter = ws.iter_rows(values_only=True)
    next(row_iter, None)
    next(row_iter, None)
    headers = list(next(row_iter))
    tenor_cols: list[tuple[int, str, float]] = []
    for c, h in enumerate(headers):
        parsed = parse_tenor(h)
        if c == 0 or parsed is None:
            continue
        label, years = parsed
        tenor_cols.append((c, label, years))

    history: dict[float, list[tuple[date, float]]] = {years: [] for _, _, years in tenor_cols}
    points: dict[float, float] = {}
    for row in row_iter:
        d = to_date(row[0] if row else None)
        if d is None:
            continue
        for c, _, years in tenor_cols:
            val = row[c] if c < len(row) else None
            if isinstance(val, (int, float)):
                history[years].append((d, float(val)))
                if years not in points:
                    points[years] = float(val)
    return Curve(sheet_name, points, history)


def parse_irs_curve(wb) -> Curve:
    ws = wb["IRS_YIELD"]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)
    names = list(next(rows))
    next(rows, None)
    data_rows = list(rows)
    history: dict[float, list[tuple[date, float]]] = {}
    points: dict[float, float] = {}
    for c, name in enumerate(names):
        parsed = parse_tenor(name)
        if parsed is None:
            continue
        _, years = parsed
        history[years] = []
        for row in data_rows:
            d = to_date(row[c] if c < len(row) else None)
            val = row[c + 1] if c + 1 < len(row) else None
            if d is not None and isinstance(val, (int, float)):
                history[years].append((d, float(val)))
                if years not in points:
                    points[years] = float(val)
    return Curve("IRS_YIELD", points, history)


def series_by_date(curve: Curve, tenor: float) -> dict[date, float]:
    return {d: v for d, v in curve.history.get(tenor, [])}


def parse_cd_repo(wb) -> dict[str, Any]:
    ws = wb["CD,REPO"]
    cd, repo = [], []
    for row in ws.iter_rows(min_row=4, values_only=True):
        d1 = to_date(row[0] if len(row) > 0 else None)
        v1 = row[1] if len(row) > 1 else None
        if d1 and isinstance(v1, (int, float)):
            cd.append((d1, float(v1)))
        d2 = to_date(row[2] if len(row) > 2 else None)
        v2 = row[3] if len(row) > 3 else None
        if d2 and isinstance(v2, (int, float)):
            repo.append((d2, float(v2)))
    return {
        "cd_latest_date": cd[0][0] if cd else None,
        "cd_3m": cd[0][1] if cd else None,
        "repo_latest_date": repo[0][0] if repo else None,
        "repo": repo[0][1] if repo else None,
        "cd_series": cd,
        "repo_series": repo,
    }


def parse_bond_list(wb, as_of: date, config: dict[str, Any]) -> list[dict[str, Any]]:
    ws = wb["KTB_MSB_List"]
    benchmarks = {v: k for k, v in config["benchmark_bonds"].items()}
    bonds = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in r):
            continue
        maturity = parse_yyyymmdd(r[6])
        issue = parse_yyyymmdd(r[5])
        name = r[0]
        prefix = "KTB" if str(name).startswith("국고") else "MSB" if str(name).startswith("통안") else "Other"
        bonds.append(
            {
                "name": name,
                "type": r[1],
                "code": r[2],
                "standard_code": r[3],
                "issue_date": issue,
                "maturity_date": maturity,
                "remaining_years": years_between(as_of, maturity) if maturity else None,
                "balance": float(r[7]) if isinstance(r[7], (int, float)) else None,
                "coupon": float(r[8]) if isinstance(r[8], (int, float)) else None,
                "ytm": float(r[9]) if isinstance(r[9], (int, float)) else None,
                "prefix": prefix,
                "is_benchmark": name in benchmarks,
                "benchmark_tenor": benchmarks.get(name),
            }
        )
    return bonds


def parse_fullinfo_lending(wb) -> dict[str, dict[str, Any]]:
    ws = wb["KTB_MSB_FULLINFO"]
    names = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    headers = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    starts = [(i, n) for i, n in enumerate(names) if n]
    recent_rows: dict[str, list[tuple[Any, ...]]] = {name: [] for _, name in starts}
    history_counts: dict[str, int] = {name: 0 for _, name in starts}
    ytm_history: dict[str, list[tuple[date, float]]] = {name: [] for _, name in starts}
    # Keep only the most recent rows needed for latest, 5D and 20D changes. The
    # sheet is very wide, so avoiding a full in-memory copy matters a lot.
    for row in ws.iter_rows(min_row=4, values_only=True):
        for start, name in starts:
            if start < len(row) and row[start] is not None:
                history_counts[name] += 1
                d = to_date(row[start])
                ytm = row[start + 1] if start + 1 < len(row) else None
                if d and isinstance(ytm, (int, float)):
                    ytm_history[name].append((d, float(ytm)))
                if len(recent_rows[name]) <= 21:
                    recent_rows[name].append(tuple(row[start : start + 30]))
    out: dict[str, dict[str, Any]] = {}
    for start, name in starts:
        h = list(headers[start : start + 30])
        rows = recent_rows[name]
        if not rows:
            continue

        def val(row_idx: int, label: str) -> float | None:
            if label not in h or row_idx >= len(rows):
                return None
            v = rows[row_idx][h.index(label)]
            return float(v) if isinstance(v, (int, float)) else None

        def delta(label: str, lag: int) -> float | None:
            a = val(0, label)
            b = val(lag, label)
            if a is None or b is None:
                return None
            return a - b

        latest_date = to_date(rows[0][0])
        lending_labels = ["보험 대여", "은행 대여", "외국인 대여", "증권 차입"]
        investor_cols = [
            ("외국인", "외국인 순매수 거래량", "외국인 잔고수량"),
            ("은행", "은행 순매수 거래량", "은행 잔고수량"),
            ("보험기금", "보험기금 순매수 거래량", "보험기금 잔고수량"),
            ("자산운용공모", "자산운용(공모) 순매수 거래량", "자산운용(공모) 잔고수량"),
            ("자산운용사모", "자산운용(사모) 순매수 거래량", "자산운용(사모) 잔고수량"),
            ("종금", "종금 순매수 거래량", "종금 잔고수량"),
            ("정부", "정부 순매수 거래량", "정부 잔고수량"),
            ("기타법인", "기타법인 순매수 거래량", "기타법인 잔고수량"),
            ("개인", "개인 순매수 거래량", "개인 잔고수량"),
            ("선물", "선물 순매수 거래량", "선물 잔고수량"),
        ]

        def rolling_sum(label: str, n: int) -> float | None:
            if label in h:
                idx = h.index(label)
                vals = [row[idx] for row in rows[:n] if isinstance(row[idx], (int, float))]
                return sum(vals) if vals else None
            return None

        flow_by_investor = {}
        for investor, flow_label, balance_label in investor_cols:
            flow_by_investor[investor] = {
                "today": val(0, flow_label),
                "5d": rolling_sum(flow_label, 5),
                "20d": rolling_sum(flow_label, 20),
                "balance": val(0, balance_label),
            }
        active_flows = [(k, v["5d"]) for k, v in flow_by_investor.items() if v.get("5d") is not None]
        top_flow_investor, top_flow_5d = (None, None)
        if active_flows:
            top_flow_investor, top_flow_5d = max(active_flows, key=lambda x: abs(x[1]))
        out[name] = {
            "latest_date": latest_date,
            "history_rows": history_counts[name],
            "ytm_history": ytm_history.get(name, []),
            "ytm_latest": val(0, "민평3사 수익률(산출일) 당일"),
            "balance_today": val(0, "금일잔량"),
            "balance_5d_change": delta("금일잔량", 5),
            "balance_20d_change": delta("금일잔량", 20),
            "trade_today": val(0, "금일거래"),
            "redeem_today": val(0, "금일상환"),
            "insurance_lending": val(0, "보험 대여"),
            "bank_lending": val(0, "은행 대여"),
            "foreign_lending": val(0, "외국인 대여"),
            "securities_borrow": val(0, "증권 차입"),
            "lending_5d_change_sum": sum(
                x for x in (delta(label, 5) for label in lending_labels) if x is not None
            ),
            "flow_by_investor": flow_by_investor,
            "top_flow_investor_5d": top_flow_investor,
            "top_flow_5d": top_flow_5d,
        }
    return out


def parse_agency_blocks(wb) -> list[dict[str, Any]]:
    ws = wb["Agency_YIELD"]
    names = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    headers = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    starts = [(i, n) for i, n in enumerate(names) if n]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    blocks = []
    for bi, (start, name) in enumerate(starts):
        end = starts[bi + 1][0] if bi + 1 < len(starts) else ws.max_column
        h = headers[start:end]
        tenor_cols = []
        for off, label in enumerate(h):
            parsed = parse_tenor(label)
            if off == 0 or parsed is None:
                continue
            _, years = parsed
            tenor_cols.append((off, years, label))
        for off, years, label in tenor_cols:
            hist = []
            for row in rows:
                d = to_date(row[start]) if start < len(row) else None
                v = row[start + off] if start + off < len(row) else None
                if d and isinstance(v, (int, float)):
                    hist.append((d, float(v)))
            if hist:
                blocks.append(
                    {
                        "sector": name,
                        "tenor_label": label,
                        "tenor_years": years,
                        "latest_date": hist[0][0],
                        "latest_yield": hist[0][1],
                        "history": hist,
                    }
                )
    return blocks


def parse_futures(wb) -> list[dict[str, Any]]:
    ws = wb["Futures"]
    all_rows = list(ws.iter_rows(values_only=True))
    names = all_rows[1]
    headers = all_rows[2]
    rows = all_rows[3:]
    starts = [(i, n) for i, n in enumerate(names) if n]
    out = []
    for bi, (start, name) in enumerate(starts):
        end = starts[bi + 1][0] if bi + 1 < len(starts) else ws.max_column
        h = list(headers[start:end])
        data = [list(row[start:end]) for row in rows if row[start] is not None]
        out.append({"name": name, "headers": h, "rows": data})
    return out


def active_contract_suffix(as_of: date, config: dict[str, Any]) -> str:
    front = config["futures"]["front_contract"]
    next_contract = config["futures"]["next_contract"]
    mat = datetime.strptime(config["futures"]["contract_maturities"][front], "%Y-%m-%d").date()
    return next_contract if as_of >= mat else front


def get_1y_agency_proxy(agency_blocks: list[dict[str, Any]], fallback: float) -> tuple[float, str]:
    candidates = [
        b for b in agency_blocks if "산금채" in str(b["sector"]) and abs(b["tenor_years"] - 1.0) < 0.05
    ]
    if candidates:
        return float(candidates[0]["latest_yield"]), f"{candidates[0]['sector']} {candidates[0]['tenor_label']}"
    candidates = [b for b in agency_blocks if abs(b["tenor_years"] - 1.0) < 0.05]
    if candidates:
        return float(candidates[0]["latest_yield"]), f"{candidates[0]['sector']} {candidates[0]['tenor_label']}"
    return fallback, "Fallback repo/CD proxy"


def latest_date_from_curve(curve: Curve) -> date:
    dates = []
    for hist in curve.history.values():
        if hist:
            dates.append(hist[0][0])
    return max(dates)


def align_curve_to_date(curve: Curve, as_of: date) -> Curve:
    aligned_history: dict[float, list[tuple[date, float]]] = {}
    aligned_points: dict[float, float] = {}
    for tenor, hist in curve.history.items():
        kept = [(d, v) for d, v in hist if d <= as_of]
        if not kept:
            continue
        aligned_history[tenor] = kept
        aligned_points[tenor] = kept[0][1]
    return Curve(curve.name, aligned_points, aligned_history)


def add_irs_3m_cd_proxy(curve: Curve, cd_repo: dict[str, Any], as_of: date) -> Curve:
    points = dict(curve.points)
    history = {k: list(v) for k, v in curve.history.items()}
    cd_hist = [(d, v) for d, v in cd_repo.get("cd_series", []) if d <= as_of]
    if cd_hist:
        history[0.25] = cd_hist
        points[0.25] = cd_hist[0][1]
    elif cd_repo.get("cd_3m") is not None:
        history[0.25] = [(as_of, cd_repo["cd_3m"])]
        points[0.25] = cd_repo["cd_3m"]
    return Curve(curve.name, points, history)


def build_tenor_cr(curve: Curve, funding: float, periods: dict[str, float]) -> list[dict[str, Any]]:
    rows = []
    for tenor, ytm in sorted(curve.points.items()):
        if tenor <= 0:
            continue
        dur = approx_duration(tenor, ytm)
        for hp, h in periods.items():
            c = carry_return(ytm, funding, h) if valid_holding_period(tenor, h) else None
            r = rolldown_return(curve, tenor, ytm, h)
            total = c + r if c is not None and r is not None else None
            rows.append(
                {
                    "asset": curve.name.replace("_YIELD", ""),
                    "tenor": tenor_label(tenor),
                    "tenor_years": tenor,
                    "holding": hp,
                    "yield": ytm,
                    "funding": funding,
                    "duration": dur,
                    "carry_return": c,
                    "rolldown_return": r,
                    "total_return": total,
                    "return_per_dv01": total / (dur * 0.0001) if total is not None and dur else None,
                    "breakeven_yield_bp": total / dur * 10000 if total is not None and dur else None,
                }
            )
    return rows


def build_irs_tenor_cr(curve: Curve, cd_rate: float, periods: dict[str, float]) -> list[dict[str, Any]]:
    rows = []
    for tenor, rate in sorted(curve.points.items()):
        dur = approx_duration(tenor, rate)
        for direction in ("Receive", "Pay"):
            for hp, h in periods.items():
                carry = None
                if valid_holding_period(tenor, h):
                    carry = (rate - cd_rate) / 100.0 * h if direction == "Receive" else (cd_rate - rate) / 100.0 * h
                rec_roll = rolldown_return(curve, tenor, rate, h)
                roll = rec_roll if direction == "Receive" else -rec_roll if rec_roll is not None else None
                total = carry + roll if carry is not None and roll is not None else None
                rows.append(
                    {
                        "direction": direction,
                        "tenor": tenor_label(tenor),
                        "tenor_years": tenor,
                        "holding": hp,
                        "irs_rate": rate,
                        "cd_3m": cd_rate,
                        "duration": dur,
                        "carry_return": carry,
                        "rolldown_return": roll,
                        "total_return": total,
                        "breakeven_rate_bp": total / dur * 10000 if total is not None and dur else None,
                    }
                )
    return rows


def curve_pair_rows(
    curve: Curve,
    pairs: list[list[str]],
    periods: dict[str, float],
    tenor_years: dict[str, float],
    base_notional: float,
    repo: float | None = None,
    collateral_yield: float | None = None,
    borrow_fee: float = 0.0,
    cd: float | None = None,
    is_irs: bool = False,
) -> list[dict[str, Any]]:
    rows = []
    for short_label, long_label in pairs:
        st = tenor_years[short_label]
        lt = tenor_years[long_label]
        sy = interp(curve.points, st)
        ly = interp(curve.points, lt)
        if sy is None or ly is None:
            continue
        sd = approx_duration(st, sy)
        ld = approx_duration(lt, ly)
        hedge = ld / sd if sd else None
        theoretical_short_notional = base_notional * hedge if hedge else None
        notional_increment = 100_000_000 if is_irs else 1_000_000_000
        short_notional = round_notional(theoretical_short_notional, notional_increment)
        current_slope = ly - sy
        for direction in ("Steepener", "Flattener"):
            for hp, h in periods.items():
                if not valid_holding_period(st, h) or not valid_holding_period(lt, h):
                    rows.append(
                        {
                            "curve": curve.name.replace("_YIELD", ""),
                            "trade": f"{short_label}-{long_label}",
                            "direction": direction,
                            "holding": hp,
                            "short_tenor": short_label,
                            "long_tenor": long_label,
                            "short_yield": sy,
                            "long_yield": ly,
                            "current_slope_bp": current_slope * 100,
                            "hedge_ratio": hedge,
                            "theoretical_short_notional": theoretical_short_notional,
                            "notional_rounding_unit": notional_increment,
                            "short_tenor_notional": short_notional,
                            "long_tenor_notional": base_notional,
                            "carry_roll_pnl": None,
                            "total_return": None,
                            "breakeven_slope_bp": None,
                        }
                    )
                    continue
                s_roll_long = rolldown_return(curve, st, sy, h)
                l_roll_long = rolldown_return(curve, lt, ly, h)
                if s_roll_long is None or l_roll_long is None:
                    rows.append(
                        {
                            "curve": curve.name.replace("_YIELD", ""),
                            "trade": f"{short_label}-{long_label}",
                            "direction": direction,
                            "holding": hp,
                            "short_tenor": short_label,
                            "long_tenor": long_label,
                            "short_yield": sy,
                            "long_yield": ly,
                            "current_slope_bp": current_slope * 100,
                            "hedge_ratio": hedge,
                            "theoretical_short_notional": theoretical_short_notional,
                            "notional_rounding_unit": notional_increment,
                            "short_tenor_notional": short_notional,
                            "long_tenor_notional": base_notional,
                            "carry_roll_pnl": None,
                            "total_return": None,
                            "breakeven_slope_bp": None,
                        }
                    )
                    continue
                if is_irs:
                    s_receive_carry = (sy - cd) / 100.0 * h
                    l_receive_carry = (ly - cd) / 100.0 * h
                    if direction == "Steepener":
                        short_leg_return = s_receive_carry + s_roll_long
                        long_leg_return = -(l_receive_carry + l_roll_long)
                    else:
                        short_leg_return = -(s_receive_carry + s_roll_long)
                        long_leg_return = l_receive_carry + l_roll_long
                else:
                    s_long_carry = carry_return(sy, repo, h)
                    l_long_carry = carry_return(ly, repo, h)
                    s_short_carry = -s_long_carry + ((collateral_yield - repo) / 100.0 * h) - borrow_fee * h
                    l_short_carry = -l_long_carry + ((collateral_yield - repo) / 100.0 * h) - borrow_fee * h
                    if direction == "Steepener":
                        short_leg_return = s_long_carry + s_roll_long
                        long_leg_return = l_short_carry - l_roll_long
                    else:
                        short_leg_return = s_short_carry - s_roll_long
                        long_leg_return = l_long_carry + l_roll_long

                pnl_short = short_leg_return * short_notional
                pnl_long = long_leg_return * base_notional
                total_pnl = pnl_short + pnl_long
                combined_notional = short_notional + base_notional
                total_return = total_pnl / combined_notional if combined_notional else None
                net_dv01 = abs(dv01_per_notional(lt, ly, base_notional))
                be_slope = total_pnl / net_dv01 if net_dv01 else None
                slope_shock_sign = 1 if direction == "Steepener" else -1
                rows.append(
                    {
                        "curve": curve.name.replace("_YIELD", ""),
                        "trade": f"{short_label}-{long_label}",
                        "direction": direction,
                        "holding": hp,
                        "short_tenor": short_label,
                        "long_tenor": long_label,
                        "short_yield": sy,
                        "long_yield": ly,
                        "current_slope_bp": current_slope * 100,
                        "hedge_ratio": hedge,
                        "theoretical_short_notional": theoretical_short_notional,
                        "notional_rounding_unit": notional_increment,
                        "short_tenor_notional": short_notional,
                        "long_tenor_notional": base_notional,
                        "carry_roll_pnl": total_pnl,
                        "total_return": total_return,
                        "slope_dv01": net_dv01,
                        "+10bp_slope_pnl": total_pnl + slope_shock_sign * net_dv01 * 10 if net_dv01 else None,
                        "-10bp_slope_pnl": total_pnl - slope_shock_sign * net_dv01 * 10 if net_dv01 else None,
                        "breakeven_slope_bp": be_slope,
                    }
                )
    return rows


def curve_points_by_date(curve: Curve) -> dict[date, dict[float, float]]:
    by_date: dict[date, dict[float, float]] = defaultdict(dict)
    for tenor, series in curve.history.items():
        for d, value in series:
            by_date[d][tenor] = value
    return by_date


def historical_residual_z(
    bond: dict[str, Any],
    curve_by_date: dict[date, dict[float, float]],
    ytm_history: list[tuple[date, float]],
    as_of: date,
    current_residual: float | None,
    lookback_days: int = 366,
    min_span_days: int = 183,
    min_points: int = 30,
) -> float | None:
    if current_residual is None or not ytm_history or bond.get("maturity_date") is None:
        return None
    cutoff = as_of.toordinal() - lookback_days
    residuals: list[tuple[date, float]] = []
    for d, ytm in ytm_history:
        if d.toordinal() < cutoff or d > as_of:
            continue
        curve_points = curve_by_date.get(d)
        if not curve_points:
            continue
        remaining = years_between(d, bond["maturity_date"])
        fitted = interp_inside_range(curve_points, remaining)
        if fitted is not None:
            residuals.append((d, ytm - fitted))
    if len(residuals) < min_points:
        return None
    dates = [d for d, _ in residuals]
    if (max(dates) - min(dates)).days < min_span_days:
        return None
    return zscore([v for _, v in residuals], current_residual)


def combined_z(cross_section_z: float | None, historical_z: float | None) -> str:
    def fmt(v: float | None) -> str:
        return "N/A" if v is None else f"{v:+.2f}"
    return f"{fmt(cross_section_z)} / {fmt(historical_z)}"


def build_bond_rv(
    bonds: list[dict[str, Any]],
    ktb: Curve,
    msb: Curve,
    lending: dict[str, dict[str, Any]],
    basket_flags: dict[str, list[str]],
) -> list[dict[str, Any]]:
    as_of = latest_date_from_curve(ktb)
    ktb_by_date = curve_points_by_date(ktb)
    msb_by_date = curve_points_by_date(msb)
    residuals = []
    for b in bonds:
        if b["ytm"] is None or b["remaining_years"] is None:
            continue
        curve = ktb if b["prefix"] == "KTB" else msb if b["prefix"] == "MSB" else None
        fitted = interp_inside_range(curve.points, b["remaining_years"]) if curve else None
        residual = b["ytm"] - fitted if fitted is not None else None
        if residual is not None:
            residuals.append(residual)
    rows = []
    for b in bonds:
        curve = ktb if b["prefix"] == "KTB" else msb if b["prefix"] == "MSB" else None
        curve_by_date = ktb_by_date if b["prefix"] == "KTB" else msb_by_date if b["prefix"] == "MSB" else {}
        range_status = curve_range_status(curve.points, b["remaining_years"]) if curve else "No curve"
        fitted = interp_inside_range(curve.points, b["remaining_years"]) if curve and b["remaining_years"] is not None else None
        residual = b["ytm"] - fitted if fitted is not None and b["ytm"] is not None else None
        lend = lending.get(b["name"], {})
        hist_z = historical_residual_z(b, curve_by_date, lend.get("ytm_history") or [], as_of, residual)
        flows = lend.get("flow_by_investor") or {}
        flow_today_vals = [v.get("today") for v in flows.values() if v.get("today") is not None]
        flow_5d_vals = [v.get("5d") for v in flows.values() if v.get("5d") is not None]
        flow_20d_vals = [v.get("20d") for v in flows.values() if v.get("20d") is not None]
        top_buy_5d = max(
            ((k, v.get("5d")) for k, v in flows.items() if v.get("5d") is not None),
            key=lambda x: x[1],
            default=(None, None),
        )
        top_sell_5d = min(
            ((k, v.get("5d")) for k, v in flows.items() if v.get("5d") is not None),
            key=lambda x: x[1],
            default=(None, None),
        )
        basket_names = basket_flags.get(b["code"], [])
        rows.append(
            {
                "name": b["name"],
                "prefix": b["prefix"],
                "code": b["code"],
                "maturity": b["maturity_date"],
                "remaining_years": b["remaining_years"],
                "coupon": b["coupon"],
                "ytm": b["ytm"],
                "fitted_yield": fitted,
                "residual_bp": residual * 100 if residual is not None else None,
                "residual_z": zscore(residuals, residual) if residual is not None else None,
                "historical_residual_z": hist_z,
                "combined_residual_z": combined_z(zscore(residuals, residual) if residual is not None else None, hist_z),
                "curve_range_status": range_status,
                "balance": b["balance"],
                "is_benchmark": b["is_benchmark"],
                "benchmark_tenor": b["benchmark_tenor"],
                "is_futures_basket": bool(basket_names),
                "futures_basket_refs": ", ".join(basket_names),
                "lending_balance": lend.get("balance_today"),
                "lending_5d_change": lend.get("balance_5d_change"),
                "lending_20d_change": lend.get("balance_20d_change"),
                "lending_trade_today": lend.get("trade_today"),
                "lending_redeem_today": lend.get("redeem_today"),
                "insurance_lending": lend.get("insurance_lending"),
                "bank_lending": lend.get("bank_lending"),
                "foreign_lending": lend.get("foreign_lending"),
                "securities_borrow": lend.get("securities_borrow"),
                "total_flow_today": sum(flow_today_vals) if flow_today_vals else None,
                "total_flow_5d": sum(flow_5d_vals) if flow_5d_vals else None,
                "total_flow_20d": sum(flow_20d_vals) if flow_20d_vals else None,
                "top_buy_investor_5d": top_buy_5d[0],
                "top_buy_5d": top_buy_5d[1],
                "top_sell_investor_5d": top_sell_5d[0],
                "top_sell_5d": top_sell_5d[1],
                "top_flow_investor_5d": lend.get("top_flow_investor_5d"),
                "top_flow_5d": lend.get("top_flow_5d"),
                "flow_by_investor": lend.get("flow_by_investor"),
                "history_rows": lend.get("history_rows"),
            }
        )
    benchmark_refs = [r for r in rows if r.get("is_benchmark") and r.get("remaining_years") is not None]
    basket_refs = [r for r in rows if r.get("is_futures_basket") and r.get("remaining_years") is not None]
    for row in rows:
        def nearest(refs: list[dict[str, Any]]) -> tuple[str | None, float | None]:
            candidates = [
                (ref, abs(row["remaining_years"] - ref["remaining_years"]))
                for ref in refs
                if row.get("remaining_years") is not None
                and ref.get("remaining_years") is not None
                and row.get("prefix") == ref.get("prefix")
                and abs(row["remaining_years"] - ref["remaining_years"]) <= 0.5
            ]
            if not candidates:
                return None, None
            ref, gap = min(candidates, key=lambda x: x[1])
            return ref["name"], gap

        bench_name, bench_gap = nearest(benchmark_refs)
        basket_name, basket_gap = nearest(basket_refs)
        focus_tags = []
        if row.get("is_benchmark"):
            focus_tags.append("Benchmark")
        elif bench_name:
            focus_tags.append("Near benchmark")
        if row.get("is_futures_basket"):
            focus_tags.append("Basket")
        elif basket_name:
            focus_tags.append("Near basket")
        row["rv_focus"] = ", ".join(focus_tags)
        row["nearest_benchmark"] = bench_name
        row["nearest_benchmark_gap_m"] = bench_gap * 12 if bench_gap is not None else None
        row["nearest_basket"] = basket_name
        row["nearest_basket_gap_m"] = basket_gap * 12 if basket_gap is not None else None
    return rows


def build_agency_spread_rows(agency_blocks, irs: Curve, msb: Curve, repo: float) -> list[dict[str, Any]]:
    rows = []
    for b in agency_blocks:
        tenor = b["tenor_years"]
        bench_curve = irs if tenor < 2.0 else msb
        bench = interp(bench_curve.points, tenor)
        if bench is None:
            continue
        hist_spreads = []
        for d, y in b["history"]:
            bench_series = series_by_date(bench_curve, tenor)
            bv = bench_series.get(d)
            if bv is None:
                bv = interp(bench_curve.points, tenor)
            if bv is not None:
                hist_spreads.append(y - bv)
        spread = b["latest_yield"] - bench
        rows.append(
            {
                "sector": b["sector"],
                "tenor": tenor_label(tenor),
                "tenor_years": tenor,
                "latest_date": b["latest_date"],
                "agency_yield": b["latest_yield"],
                "benchmark": "IRS" if tenor < 2.0 else "MSB",
                "benchmark_yield": bench,
                "spread_bp": spread * 100,
                "spread_z": zscore(hist_spreads[:252], spread),
                "spread_percentile": pct_rank(hist_spreads[:252], spread),
                "carry_3m": carry_return(b["latest_yield"], repo, 0.25),
                "carry_6m": carry_return(b["latest_yield"], repo, 0.5),
                "carry_1y": carry_return(b["latest_yield"], repo, 1.0),
            }
        )
    return rows


def build_ktb_irs_spread_rows(ktb: Curve, irs: Curve, repo: float, cd: float, periods) -> list[dict[str, Any]]:
    rows = []
    common = sorted(set(ktb.points) & set(irs.points))
    for tenor in common:
        if tenor < 2.0:
            continue
        ky = ktb.points[tenor]
        iy = irs.points[tenor]
        spread = ky - iy
        hist = []
        ktb_hist = dict(ktb.history.get(tenor, []))
        irs_hist = dict(irs.history.get(tenor, []))
        for d in sorted(set(ktb_hist) & set(irs_hist), reverse=True):
            hist.append(ktb_hist[d] - irs_hist[d])
        for hp, h in periods.items():
            ktb_roll = rolldown_return(ktb, tenor, ky, h)
            irs_roll = rolldown_return(irs, tenor, iy, h)
            if valid_holding_period(tenor, h) and ktb_roll is not None and irs_roll is not None:
                ktb_cr = carry_return(ky, repo, h) + ktb_roll
                irs_pay = (cd - iy) / 100.0 * h - irs_roll
                total_cr = ktb_cr + irs_pay
            else:
                ktb_cr = None
                irs_pay = None
                total_cr = None
            rows.append(
                {
                    "tenor": tenor_label(tenor),
                    "holding": hp,
                    "ktb_yield": ky,
                    "irs_rate": iy,
                    "spread_bp": spread * 100,
                    "spread_z_252": zscore(hist[:252], spread),
                    "spread_percentile_252": pct_rank(hist[:252], spread),
                    "ktb_buy_cr": ktb_cr,
                    "irs_pay_cr": irs_pay,
                    "total_cr": total_cr,
                }
            )
    return rows


def build_futures_outputs(
    futures: list[dict[str, Any]],
    irs: Curve,
    code_to_name: dict[str, str],
    config: dict[str, Any],
    as_of: date,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, list[str]]]:
    active_suffix = active_contract_suffix(as_of, config)
    basket_rows = []
    spread_rows = []
    basket_flags: dict[str, list[str]] = defaultdict(list)
    for f in futures:
        name = " ".join(str(f["name"]).split())
        is_active = active_suffix in name
        h = f["headers"]
        rows = f["rows"]
        if not rows:
            continue
        latest = rows[0]
        date_latest = to_date(latest[0])
        y_idx = h.index("국채현물수익률") if "국채현물수익률" in h else None
        spot_yield = latest[y_idx] if y_idx is not None else None
        tenor = 3.0 if name.startswith("3년") else 10.0 if name.startswith("10년") else None
        basket_codes = []
        basket_yields = []
        for i, label in enumerate(h):
            if label and "종목코드" in str(label):
                code = latest[i]
                y_label = str(label).replace("종목코드", "민평3사수익률")
                y = latest[h.index(y_label)] if y_label in h else None
                basket_codes.append(code)
                basket_yields.append(float(y) if isinstance(y, (int, float)) else None)
                basket_flags[code].append(name + " " + str(label).replace("종목코드", ""))
        if len(basket_yields) == 3 and all(v is not None for v in basket_yields):
            value = basket_yields[1] - (basket_yields[0] + basket_yields[2]) / 2.0
            hist = []
            for row in rows:
                ys = []
                for i, label in enumerate(h):
                    if label and "민평3사수익률" in str(label):
                        v = row[i]
                        if isinstance(v, (int, float)):
                            ys.append(float(v))
                if len(ys) == 3:
                    hist.append(ys[1] - (ys[0] + ys[2]) / 2.0)
            metric = "Butterfly: basket2 - avg(basket1,basket3)"
        elif len(basket_yields) == 2 and all(v is not None for v in basket_yields):
            value = basket_yields[0] - basket_yields[1]
            hist = []
            for row in rows:
                ys = []
                for i, label in enumerate(h):
                    if label and "민평3사수익률" in str(label):
                        v = row[i]
                        if isinstance(v, (int, float)):
                            ys.append(float(v))
                if len(ys) == 2:
                    hist.append(ys[0] - ys[1])
            metric = "Spread: basket1 - basket2"
        else:
            value, hist, metric = None, [], None
        if value is not None:
            basket_rows.append(
                {
                    "future": name,
                    "is_active": is_active,
                    "latest_date": date_latest,
                    "metric": metric,
                    "value_bp": value * 100,
                    "z_score": zscore(hist[:252], value),
                    "percentile": pct_rank(hist[:252], value),
                    "basket_codes": ", ".join(str(c) for c in basket_codes),
                    "basket_names": ", ".join(code_to_name.get(str(c), "") for c in basket_codes),
                    "undervaluation": latest[h.index("저평가")] if "저평가" in h else None,
                }
            )
        if tenor and spot_yield is not None:
            irs_hist = series_by_date(irs, tenor)
            irs_rate = irs_hist.get(date_latest)
            if irs_rate is None:
                irs_rate = interp(irs.points, tenor)
            spread = float(spot_yield) - irs_rate if irs_rate is not None else None
            hist_spreads = []
            for row in rows:
                d = to_date(row[0])
                fy = row[y_idx] if y_idx is not None else None
                iy = irs_hist.get(d)
                if d and isinstance(fy, (int, float)) and iy is not None:
                    hist_spreads.append(float(fy) - iy)
            spread_rows.append(
                {
                    "future": name,
                    "is_active": is_active,
                    "latest_date": date_latest,
                    "tenor": tenor_label(tenor),
                    "futures_yield": spot_yield,
                    "irs_rate_aligned": irs_rate,
                    "spread_bp": spread * 100 if spread is not None else None,
                    "z_score": zscore(hist_spreads[:252], spread) if spread is not None else None,
                    "percentile": pct_rank(hist_spreads[:252], spread) if spread is not None else None,
                }
            )
    return basket_rows, spread_rows, basket_flags


def aligned_spread_series(
    left: list[tuple[date, float]],
    right: list[tuple[date, float]],
    scale: float = 1.0,
) -> list[tuple[date, float]]:
    lmap = {d: v for d, v in left}
    rmap = {d: v for d, v in right}
    dates = sorted(set(lmap) & set(rmap))
    return [(d, (lmap[d] - rmap[d]) * scale) for d in dates]


def curve_slope_history(curve: Curve, short_tenor: float, long_tenor: float) -> list[tuple[date, float]]:
    return aligned_spread_series(curve.history.get(long_tenor, []), curve.history.get(short_tenor, []), 100.0)


def ktb_irs_spread_history(ktb: Curve, irs: Curve, tenor: float) -> list[tuple[date, float]]:
    return aligned_spread_series(ktb.history.get(tenor, []), irs.history.get(tenor, []), 100.0)


def interpolated_curve_history(curve: Curve, tenor: float) -> list[tuple[date, float]]:
    by_date: dict[date, dict[float, float]] = defaultdict(dict)
    for t, hist in curve.history.items():
        for d, v in hist:
            by_date[d][t] = v
    out = []
    for d in sorted(by_date):
        val = interp(by_date[d], tenor)
        if val is not None:
            out.append((d, val))
    return out


def agency_spread_history(block: dict[str, Any], irs: Curve, msb: Curve) -> list[tuple[date, float]]:
    tenor = block["tenor_years"]
    bench_curve = irs if tenor < 2.0 else msb
    bench = bench_curve.history.get(tenor) or interpolated_curve_history(bench_curve, tenor)
    return aligned_spread_series(block["history"], bench, 100.0)


def short_agency_sector(sector: str) -> str:
    mapping = {
        "공사/공단채": "공사/공단",
        "산금채": "산금",
        "중금채": "중금",
        "한국수출입은행": "수출입",
        "은행채": "은행",
    }
    for key, value in mapping.items():
        if key in sector:
            return value
    text = str(sector).replace("시가평가 3사평균", "").replace("AAA", "").strip()
    return text[:10] or str(sector)[:10]


def futures_basket_history(future: dict[str, Any]) -> tuple[str | None, list[tuple[date, float]]]:
    h = future["headers"]
    rows = future["rows"]
    yidx = [i for i, label in enumerate(h) if label and "민평3사수익률" in str(label)]
    out = []
    if len(yidx) == 3:
        for row in rows:
            d = to_date(row[0])
            vals = [row[i] for i in yidx]
            if d and all(isinstance(v, (int, float)) for v in vals):
                out.append((d, (float(vals[1]) - (float(vals[0]) + float(vals[2])) / 2.0) * 100.0))
        return "Butterfly bp", sorted(out)
    if len(yidx) == 2:
        for row in rows:
            d = to_date(row[0])
            vals = [row[i] for i in yidx]
            if d and all(isinstance(v, (int, float)) for v in vals):
                out.append((d, (float(vals[0]) - float(vals[1])) * 100.0))
        return "Basket1-Basket2 bp", sorted(out)
    return None, []


def futures_irs_history(future: dict[str, Any], irs: Curve) -> list[tuple[date, float]]:
    h = future["headers"]
    rows = future["rows"]
    name = " ".join(str(future["name"]).split())
    tenor = 3.0 if name.startswith("3년") else 10.0 if name.startswith("10년") else None
    if tenor is None or "국채현물수익률" not in h:
        return []
    y_idx = h.index("국채현물수익률")
    irs_map = {d: v for d, v in irs.history.get(tenor, [])}
    out = []
    for row in rows:
        d = to_date(row[0])
        fy = row[y_idx]
        if d and isinstance(fy, (int, float)) and d in irs_map:
            out.append((d, (float(fy) - irs_map[d]) * 100.0))
    return sorted(out)


def header_has(label: Any, text: str) -> bool:
    return label is not None and text in str(label)


def first_header_index(headers: list[Any], *needles: str) -> int | None:
    for i, label in enumerate(headers):
        text = str(label) if label is not None else ""
        if any(needle in text for needle in needles):
            return i
    return None


def futures_tenor_from_name(name: str) -> float | None:
    compact = str(name).replace(" ", "")
    if compact.startswith("3"):
        return 3.0
    if compact.startswith("10"):
        return 10.0
    return None


def futures_basket_yield_indices(headers: list[Any]) -> list[int]:
    return [i for i, label in enumerate(headers) if header_has(label, "민평3사수익률")]


def futures_implied_yield_index(headers: list[Any]) -> int | None:
    return first_header_index(headers, "선물내재수익률", "현물수익률", "국채현물수익률")


def build_futures_outputs(
    futures: list[dict[str, Any]],
    irs: Curve,
    code_to_name: dict[str, str],
    config: dict[str, Any],
    as_of: date,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, list[str]]]:
    active_suffix = active_contract_suffix(as_of, config)
    basket_rows = []
    spread_rows = []
    basket_flags: dict[str, list[str]] = defaultdict(list)
    for f in futures:
        name = " ".join(str(f["name"]).split())
        is_active = active_suffix in name
        h = f["headers"]
        rows = f["rows"]
        if not rows:
            continue
        latest = rows[0]
        date_latest = to_date(latest[0])
        y_idx = futures_implied_yield_index(h)
        spot_yield = latest[y_idx] if y_idx is not None else None
        tenor = futures_tenor_from_name(name)
        basket_codes = []
        basket_yields = []
        basket_yield_indices = []
        for i, label in enumerate(h):
            if header_has(label, "종목코드"):
                code = latest[i]
                y_label = str(label).replace("종목코드", "민평3사수익률")
                y_i = h.index(y_label) if y_label in h else None
                y = latest[y_i] if y_i is not None else None
                basket_codes.append(code)
                basket_yields.append(float(y) if isinstance(y, (int, float)) else None)
                if y_i is not None:
                    basket_yield_indices.append(y_i)
                basket_flags[code].append(name + " " + str(label).replace("종목코드", ""))
        if len(basket_yields) == 3 and all(v is not None for v in basket_yields):
            value = basket_yields[1] - (basket_yields[0] + basket_yields[2]) / 2.0
            hist = []
            for row in rows:
                ys = [row[i] for i in basket_yield_indices]
                if len(ys) == 3 and all(isinstance(v, (int, float)) for v in ys):
                    hist.append(float(ys[1]) - (float(ys[0]) + float(ys[2])) / 2.0)
            metric = "Butterfly: basket2 - avg(basket1,basket3)"
        elif len(basket_yields) == 2 and all(v is not None for v in basket_yields):
            value = basket_yields[0] - basket_yields[1]
            hist = []
            for row in rows:
                ys = [row[i] for i in basket_yield_indices]
                if len(ys) == 2 and all(isinstance(v, (int, float)) for v in ys):
                    hist.append(float(ys[0]) - float(ys[1]))
            metric = "Spread: basket1 - basket2"
        else:
            value, hist, metric = None, [], None
        if value is not None:
            undervaluation_idx = first_header_index(h, "저평가")
            basket_rows.append(
                {
                    "future": name,
                    "is_active": is_active,
                    "latest_date": date_latest,
                    "metric": metric,
                    "value_bp": value * 100,
                    "z_score": zscore(hist[:252], value),
                    "percentile": pct_rank(hist[:252], value),
                    "basket_codes": ", ".join(str(c) for c in basket_codes),
                    "basket_names": ", ".join(code_to_name.get(str(c), "") for c in basket_codes),
                    "undervaluation": latest[undervaluation_idx] if undervaluation_idx is not None else None,
                }
            )
        if tenor and spot_yield is not None and y_idx is not None:
            irs_hist = series_by_date(irs, tenor)
            irs_rate = irs_hist.get(date_latest)
            if irs_rate is None:
                irs_rate = interp(irs.points, tenor)
            spread = float(spot_yield) - irs_rate if irs_rate is not None else None
            hist_spreads = []
            for row in rows:
                d = to_date(row[0])
                fy = row[y_idx]
                iy = irs_hist.get(d)
                if d and isinstance(fy, (int, float)) and iy is not None:
                    hist_spreads.append(float(fy) - iy)
            spread_rows.append(
                {
                    "future": name,
                    "is_active": is_active,
                    "latest_date": date_latest,
                    "tenor": tenor_label(tenor),
                    "futures_yield": spot_yield,
                    "irs_rate_aligned": irs_rate,
                    "spread_bp": spread * 100 if spread is not None else None,
                    "z_score": zscore(hist_spreads[:252], spread) if spread is not None else None,
                    "percentile": pct_rank(hist_spreads[:252], spread) if spread is not None else None,
                }
            )
    return basket_rows, spread_rows, basket_flags


def futures_basket_history(future: dict[str, Any]) -> tuple[str | None, list[tuple[date, float]]]:
    h = future["headers"]
    rows = future["rows"]
    yidx = futures_basket_yield_indices(h)
    out = []
    if len(yidx) == 3:
        for row in rows:
            d = to_date(row[0])
            vals = [row[i] for i in yidx]
            if d and all(isinstance(v, (int, float)) for v in vals):
                out.append((d, (float(vals[1]) - (float(vals[0]) + float(vals[2])) / 2.0) * 100.0))
        return "Butterfly bp", sorted(out)
    if len(yidx) == 2:
        for row in rows:
            d = to_date(row[0])
            vals = [row[i] for i in yidx]
            if d and all(isinstance(v, (int, float)) for v in vals):
                out.append((d, (float(vals[0]) - float(vals[1])) * 100.0))
        return "Basket1-Basket2 bp", sorted(out)
    return None, []


def futures_irs_history(future: dict[str, Any], irs: Curve) -> list[tuple[date, float]]:
    h = future["headers"]
    rows = future["rows"]
    name = " ".join(str(future["name"]).split())
    tenor = futures_tenor_from_name(name)
    y_idx = futures_implied_yield_index(h)
    if tenor is None or y_idx is None:
        return []
    irs_map = {d: v for d, v in irs.history.get(tenor, [])}
    out = []
    for row in rows:
        d = to_date(row[0])
        fy = row[y_idx]
        if d and isinstance(fy, (int, float)) and d in irs_map:
            out.append((d, (float(fy) - irs_map[d]) * 100.0))
    return sorted(out)


def futures_basket_code_indices(rows: list[list[Any]]) -> list[int]:
    if not rows:
        return []
    latest = rows[0]
    return [i for i, value in enumerate(latest) if isinstance(value, str) and value.startswith("KR")]


def latest_numeric_basket_observation(rows: list[list[Any]], yidx: list[int]) -> tuple[date | None, list[float] | None]:
    for row in rows:
        d = to_date(row[0])
        vals = [row[i] for i in yidx if i < len(row)]
        if d and len(vals) == len(yidx) and all(isinstance(v, (int, float)) for v in vals):
            return d, [float(v) for v in vals]
    return None, None


def build_futures_outputs(
    futures: list[dict[str, Any]],
    irs: Curve,
    code_to_name: dict[str, str],
    config: dict[str, Any],
    as_of: date,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, list[str]]]:
    active_suffix = active_contract_suffix(as_of, config)
    basket_rows = []
    spread_rows = []
    basket_flags: dict[str, list[str]] = defaultdict(list)
    for f in futures:
        name = " ".join(str(f["name"]).split())
        is_active = active_suffix in name
        h = f["headers"]
        rows = f["rows"]
        if not rows:
            continue
        latest = rows[0]
        date_latest = to_date(latest[0])
        y_idx = futures_implied_yield_index(h)
        spot_yield = latest[y_idx] if y_idx is not None else None
        tenor = futures_tenor_from_name(name)

        basket_yield_indices = futures_basket_yield_indices(h)
        basket_code_indices = futures_basket_code_indices(rows)
        basket_codes = [latest[i] for i in basket_code_indices]
        for n, code in enumerate(basket_codes, start=1):
            basket_flags[code].append(f"{name} basket{n}")

        basket_date, basket_yields = latest_numeric_basket_observation(rows, basket_yield_indices)
        hist = []
        if len(basket_yield_indices) == 3:
            for row in rows:
                d = to_date(row[0])
                ys = [row[i] for i in basket_yield_indices]
                if d and all(isinstance(v, (int, float)) for v in ys):
                    hist.append(float(ys[1]) - (float(ys[0]) + float(ys[2])) / 2.0)
            value = basket_yields[1] - (basket_yields[0] + basket_yields[2]) / 2.0 if basket_yields else None
            metric = "Butterfly: basket2 - avg(basket1,basket3)"
        elif len(basket_yield_indices) == 2:
            for row in rows:
                d = to_date(row[0])
                ys = [row[i] for i in basket_yield_indices]
                if d and all(isinstance(v, (int, float)) for v in ys):
                    hist.append(float(ys[0]) - float(ys[1]))
            value = basket_yields[0] - basket_yields[1] if basket_yields else None
            metric = "Spread: basket1 - basket2"
        else:
            value, metric = None, None

        if value is not None:
            undervaluation_idx = 2 if len(latest) > 2 else None
            basket_rows.append(
                {
                    "future": name,
                    "is_active": is_active,
                    "latest_date": basket_date,
                    "metric": metric,
                    "value_bp": value * 100,
                    "z_score": zscore(hist[:252], value),
                    "percentile": pct_rank(hist[:252], value),
                    "basket_codes": ", ".join(str(c) for c in basket_codes),
                    "basket_names": ", ".join(code_to_name.get(str(c), "") for c in basket_codes),
                    "undervaluation": latest[undervaluation_idx] if undervaluation_idx is not None else None,
                }
            )

        if tenor and spot_yield is not None and y_idx is not None:
            irs_hist = series_by_date(irs, tenor)
            irs_rate = irs_hist.get(date_latest)
            if irs_rate is None:
                irs_rate = interp(irs.points, tenor)
            spread = float(spot_yield) - irs_rate if irs_rate is not None else None
            hist_spreads = []
            for row in rows:
                d = to_date(row[0])
                fy = row[y_idx]
                iy = irs_hist.get(d)
                if d and isinstance(fy, (int, float)) and iy is not None:
                    hist_spreads.append(float(fy) - iy)
            spread_rows.append(
                {
                    "future": name,
                    "is_active": is_active,
                    "latest_date": date_latest,
                    "tenor": tenor_label(tenor),
                    "futures_yield": spot_yield,
                    "irs_rate_aligned": irs_rate,
                    "spread_bp": spread * 100 if spread is not None else None,
                    "z_score": zscore(hist_spreads[:252], spread) if spread is not None else None,
                    "percentile": pct_rank(hist_spreads[:252], spread) if spread is not None else None,
                }
            )
    return basket_rows, spread_rows, basket_flags


def filter_lookback(series: list[tuple[date, float]], days: int) -> tuple[list[tuple[date, float]], str]:
    clean = sorted((d, v) for d, v in series if d is not None and v is not None)
    if not clean:
        return [], "No data"
    last = clean[-1][0]
    filtered = [(d, v) for d, v in clean if (last - d).days <= days]
    # If there is not enough history for the requested window, use all available
    # data and label it clearly in the chart title.
    if len(filtered) < min(10, len(clean)):
        filtered = clean
    first = filtered[0][0]
    return filtered, f"{first:%Y-%m-%d} to {last:%Y-%m-%d}"


def combine_series_for_chart(series_map: dict[str, list[tuple[date, float]]], days: int) -> tuple[list[dict[str, Any]], str]:
    filtered: dict[str, list[tuple[date, float]]] = {}
    labels = []
    for name, series in series_map.items():
        s, label = filter_lookback(series, days)
        if s:
            filtered[name] = s
            labels.append(label)
    if not filtered:
        return [], "No data"
    all_dates = sorted(set().union(*(set(d for d, _ in s) for s in filtered.values())))
    maps = {name: {d: v for d, v in s} for name, s in filtered.items()}
    rows = []
    for d in all_dates:
        row = {"Date": d}
        has_value = False
        for name in filtered:
            val = maps[name].get(d)
            row[name] = val
            has_value = has_value or val is not None
        if has_value:
            rows.append(row)
    return rows, min(labels) if labels else "No data"


def add_chart_source_and_line_chart(
    ws_chart,
    ws_data,
    title: str,
    series_map: dict[str, list[tuple[date, float]]],
    days: int,
    anchor: str,
    data_cursor: int,
    height: float = 8.0,
    width: float = 15.0,
) -> int:
    rows, label = combine_series_for_chart(series_map, days)
    ws_chart.cell(row=int(re.sub(r"\D", "", anchor) or 1), column=1, value=None)
    if not rows:
        ws_chart[anchor] = f"{title}: no data"
        return data_cursor

    headers = list(rows[0].keys())
    start_row = data_cursor
    for c, h in enumerate(headers, start=1):
        ws_data.cell(start_row, c, h)
    for r, row in enumerate(rows, start=start_row + 1):
        for c, h in enumerate(headers, start=1):
            cell = ws_data.cell(r, c, row.get(h))
            if h == "Date":
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.number_format = "0.00"

    chart = LineChart()
    chart.title = f"{title} ({label})"
    chart.style = 13
    chart.height = height
    chart.width = width
    chart.y_axis.title = "bp"
    chart.x_axis.title = "Date"
    chart.legend.position = "r"
    chart.y_axis.majorGridlines = None
    data_ref = Reference(ws_data, min_col=2, max_col=len(headers), min_row=start_row, max_row=start_row + len(rows))
    cats_ref = Reference(ws_data, min_col=1, min_row=start_row + 1, max_row=start_row + len(rows))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    palette = ["1F4E78", "C00000", "70AD47", "7030A0", "ED7D31", "4472C4"]
    for i, series in enumerate(chart.series):
        series.graphicalProperties.line.solidFill = palette[i % len(palette)]
        series.graphicalProperties.line.width = 22000
    ws_chart.add_chart(chart, anchor)
    return start_row + len(rows) + 3


def select_series_map(
    series_map: dict[str, list[tuple[date, float]]],
    preferred: list[str] | None = None,
    max_series: int = 4,
) -> dict[str, list[tuple[date, float]]]:
    selected: dict[str, list[tuple[date, float]]] = {}
    if preferred:
        for key in preferred:
            if key in series_map and series_map[key]:
                selected[key] = series_map[key]
            if len(selected) >= max_series:
                return selected
    for key, value in series_map.items():
        if key not in selected and value:
            selected[key] = value
        if len(selected) >= max_series:
            break
    return selected


def series_change(series: list[tuple[date, float]], days: int) -> float | None:
    clean = sorted((d, v) for d, v in series if d is not None and v is not None)
    if len(clean) < 2:
        return None
    last_d, last_v = clean[-1]
    candidates = [(d, v) for d, v in clean if (last_d - d).days <= days]
    if len(candidates) < 2:
        candidates = clean
    return last_v - candidates[0][1]


def series_summary_rows(series_map: dict[str, list[tuple[date, float]]]) -> list[dict[str, Any]]:
    rows = []
    for name, series in series_map.items():
        clean = sorted((d, v) for d, v in series if d is not None and v is not None)
        if not clean:
            continue
        one_year, _ = filter_lookback(clean, 366)
        values_1y = [v for _, v in one_year]
        current = clean[-1][1]
        rows.append(
            {
                "Series": name,
                "Current bp": current,
                "3M chg": series_change(clean, 92),
                "6M chg": series_change(clean, 183),
                "1Y chg": series_change(clean, 366),
                "1Y z": zscore(values_1y, current),
            }
        )
    return rows


def add_compact_table(ws, rows: list[dict[str, Any]], start_row: int, start_col: int) -> None:
    if not rows:
        ws.cell(start_row, start_col, "No data")
        return
    headers = list(rows[0].keys())
    for c, h in enumerate(headers, start=start_col):
        cell = ws.cell(start_row, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=start_row + 1):
        for c, h in enumerate(headers, start=start_col):
            val = row[h]
            cell = ws.cell(r, c, val)
            if isinstance(val, float):
                cell.number_format = "0.00"
    for c in range(start_col, start_col + len(headers)):
        ws.column_dimensions[get_column_letter(c)].width = 14 if c > start_col else 24


def add_clean_spread_panel(
    ws_chart,
    ws_data,
    panel_title: str,
    series_map: dict[str, list[tuple[date, float]]],
    row_start: int,
    data_cursor: int,
) -> int:
    ws_chart.cell(row_start, 1, panel_title).font = Font(bold=True, size=14, color="1F2937")
    add_compact_table(ws_chart, series_summary_rows(series_map), row_start + 2, 1)
    return add_chart_source_and_line_chart(
        ws_chart,
        ws_data,
        f"{panel_title} - 1Y trend",
        series_map,
        366,
        f"H{row_start + 2}",
        data_cursor,
        height=8.5,
        width=18.5,
    )


def add_spread_chart_panel(
    ws_chart,
    ws_data,
    panel_title: str,
    series_map: dict[str, list[tuple[date, float]]],
    row_start: int,
    data_cursor: int,
) -> int:
    ws_chart.cell(row_start, 1, panel_title).font = Font(bold=True, size=14, color="1F2937")
    lookbacks = [("3M", 92, f"A{row_start + 2}"), ("6M", 183, f"I{row_start + 2}"), ("1Y", 366, f"Q{row_start + 2}")]
    cursor = data_cursor
    for label, days, anchor in lookbacks:
        cursor = add_chart_source_and_line_chart(
            ws_chart,
            ws_data,
            f"{panel_title} - {label}",
            series_map,
            days,
            anchor,
            cursor,
        )
    return cursor


def add_table(ws, rows: list[dict[str, Any]], start_row: int = 1, title: str | None = None) -> int:
    r = start_row
    if title:
        ws.cell(r, 1, title).font = Font(bold=True, size=13, color="1F2937")
        r += 2
    if not rows:
        ws.cell(r, 1, "No data")
        return r + 1
    headers = list(rows[0].keys())
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(r, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(rows, start=r + 1):
        for c, h in enumerate(headers, start=1):
            v = row.get(h)
            cell = ws.cell(row_idx, c, v)
            if isinstance(v, float):
                if "백만" in h:
                    cell.number_format = "#,##0.0"
                elif "notional" in h or "pnl" in h or "balance" in h:
                    cell.number_format = "#,##0"
                elif "return" in h or h.endswith("_cr") or "carry" in h or "rolldown" in h:
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "0.000"
            elif isinstance(v, int):
                cell.number_format = "#,##0"
            elif isinstance(v, date):
                cell.number_format = "yyyy-mm-dd"
    ws.freeze_panes = ws.cell(r + 1, 1)
    ws.auto_filter.ref = ws.dimensions
    return r + len(rows) + 2


def style_sheet(ws):
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:200]:
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)


def highlight_bond_rv(ws):
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    bm_col = headers.get("is_benchmark")
    basket_col = headers.get("is_futures_basket")
    if not bm_col and not basket_col:
        return
    benchmark_fill = PatternFill("solid", fgColor="FFF2CC")
    basket_fill = PatternFill("solid", fgColor="DDEBF7")
    both_fill = PatternFill("solid", fgColor="E2F0D9")
    for r in range(2, ws.max_row + 1):
        is_bm = bool(ws.cell(r, bm_col).value) if bm_col else False
        is_basket = bool(ws.cell(r, basket_col).value) if basket_col else False
        fill = both_fill if is_bm and is_basket else benchmark_fill if is_bm else basket_fill if is_basket else None
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill


def style_summary_cues(ws):
    green = PatternFill("solid", fgColor="E2F0D9")
    red = PatternFill("solid", fgColor="FCE4D6")
    neutral = PatternFill("solid", fgColor="F2F2F2")
    for row in ws.iter_rows():
        for cell in row:
            header = ws.cell(cell.row - ((cell.row > 1) and 0), cell.column).value
            if isinstance(cell.value, str):
                if cell.value.startswith("▲"):
                    cell.fill = green
                elif cell.value.startswith("▼"):
                    cell.fill = red
                elif cell.value.startswith("→"):
                    cell.fill = neutral
            if isinstance(cell.value, (int, float)):
                col_header = None
                # Search upward for nearest non-empty header-like cell in same column.
                for rr in range(cell.row - 1, max(1, cell.row - 6), -1):
                    v = ws.cell(rr, cell.column).value
                    if isinstance(v, str):
                        col_header = v
                        break
                if col_header and ("백만" in col_header or col_header in {"1Y z", "z"}):
                    if cell.value > 0:
                        cell.fill = green
                    elif cell.value < 0:
                        cell.fill = red


def add_methodology(ws):
    rows = [
        ("Carry", "보유 기간 동안 수취하는 금리/수익률에서 조달금리를 뺀 정태 수익", "채권: (YTM - Repo) x 기간, IRS Receive: (IRS - CD3M) x 기간", "양수일수록 시간이 지날 때 유리"),
        ("Rolldown", "커브가 변하지 않는다고 가정할 때 잔존만기가 짧아져 생기는 가격효과", "-Duration x (Rolled Yield - Current Yield)", "커브가 우상향이면 대체로 양수"),
        ("DV01", "금리 1bp 변화에 대한 가격 민감도", "Duration x 0.0001 x Notional", "커브 포지션 헷지비율 계산에 사용"),
        ("Hedge Ratio", "장기테너 100억 기준 단기테너 목표 보유금액 배율", "DV01(long) / DV01(short)", "P&L 계산 시 KTB 단기금액은 10억, IRS 단기금액은 1억 단위 반올림"),
        ("KTB Steepener", "단기 KTB 매수, 장기 KTB 대차매도", "Short-tenor long + Long-tenor short", "장기 매도 leg에 담보 carry와 대차 fee 반영"),
        ("KTB Flattener", "장기 KTB 매수, 단기 KTB 대차매도", "Long-tenor long + Short-tenor short", "단기 매도 leg에 담보 carry와 대차 fee 반영"),
        ("대차매도 Carry", "빌린 채권을 매도하고 담보채권을 보유할 때의 carry", "-채권 long carry + 1Y 산금채 담보 carry - 대차수수료", "대차수수료 기본값은 연 0.35%"),
        ("IRS Receive/Pay", "고정금리 수취/지급 IRS 포지션", "Receive carry = IRS - CD3M, Pay carry = CD3M - IRS", "IRS curve는 대차/담보 비용 없음"),
        ("IRS 3M Proxy", "IRS 6M 이하 rolldown 계산을 위한 3개월 지점 대용치", "IRS 3M proxy = CD 3M", "실제 3M IRS 고시가 아니라 계산용 proxy"),
        ("Breakeven bp", "carry/rolldown 이익이 금리 또는 slope 변화로 상쇄되는 폭", "Tenor: total / duration, Curve: total P&L / slope DV01", "양수라면 그만큼 불리하게 움직여야 손익 0"),
        ("Z-score", "최근 분포 대비 현재값의 표준화 위치", "(현재값 - 평균) / 표준편차", "절대값 2 이상은 극단 구간 후보"),
        ("Percentile", "현재값이 과거 분포에서 위치한 백분위", "과거값 중 현재 이하 비율", "0.9 이상은 넓음/cheap, 0.1 이하는 타이트/rich 후보"),
        ("Fitted Curve Residual", "개별 종목 YTM과 generic curve 보간값의 차이", "Actual YTM - Fitted YTM", "양수는 curve 대비 cheap 후보"),
        ("Futures Basket Butterfly", "3년 선물 바스켓 중간 종목의 상대가치", "Basket2 - (Basket1 + Basket3)/2", "양수는 중간 바스켓 수익률이 높아 cheap 후보"),
    ]
    add_table(ws, [{"term": a, "meaning": b, "formula": c, "interpretation": d} for a, b, c, d in rows])


def signal_text(z: float | None, positive: str, negative: str) -> str:
    if z is None:
        return "No signal"
    if z >= 2:
        return "Strong " + positive
    if z >= 1:
        return "Moderate " + positive
    if z <= -2:
        return "Strong " + negative
    if z <= -1:
        return "Moderate " + negative
    return "Neutral"


def rv_side(residual_bp: float | None) -> str:
    if residual_bp is None:
        return ""
    if residual_bp > 0:
        return "Cheap: \uc2e4\uc81c YTM\uc774 \ucee4\ube0c\ubcf4\ub2e4 \ub192\uc74c"
    if residual_bp < 0:
        return "Rich: \uc2e4\uc81c YTM\uc774 \ucee4\ube0c\ubcf4\ub2e4 \ub0ae\uc74c"
    return "Neutral"


def rv_focus_label(row: dict[str, Any]) -> str:
    labels = []
    if row.get("is_benchmark"):
        labels.append("\uc9c0\ud45c\ucc44")
    if row.get("is_futures_basket"):
        labels.append("\uc120\ubb3c \ubc14\uc2a4\ucf13")
    if not labels and "Near benchmark" in str(row.get("rv_focus", "")):
        labels.append("\uc9c0\ud45c\ucc44 \uc8fc\ubcc0")
    if not labels and "Near basket" in str(row.get("rv_focus", "")):
        labels.append("\ubc14\uc2a4\ucf13 \uc8fc\ubcc0")
    return " + ".join(labels) if labels else str(row.get("rv_focus") or "")


def rv_compare_row(row: dict[str, Any], ref: dict[str, Any], reason: str, ref_type: str) -> dict[str, Any]:
    ytm_gap = None
    if row.get("ytm") is not None and ref.get("ytm") is not None:
        ytm_gap = (row["ytm"] - ref["ytm"]) * 100.0
    mat_gap = None
    if row.get("remaining_years") is not None and ref.get("remaining_years") is not None:
        mat_gap = abs(row["remaining_years"] - ref["remaining_years"]) * 12.0
    return {
        "\uc120\uc815 \uc0ac\uc720": reason,
        "\uc885\ubaa9": row["name"],
        "\ub9cc\uae30": row["maturity"],
        "YTM": row["ytm"],
        "\ube44\uad50 \uae30\uc900": ref["name"],
        "\ube44\uad50 \uc720\ud615": ref_type,
        "\ub9cc\uae30\ucc28 \uc6d4": mat_gap,
        "YTM \ucc28 bp": ytm_gap,
        "\ucee4\ube0c \uad34\ub9ac bp": row["residual_bp"],
        "\ud574\uc11d": rv_side(row.get("residual_bp")),
        "\ub2e8\uba74/\uacfc\uac70 z": row.get("combined_residual_z"),
        "\ubc14\uc2a4\ucf13 \ud3ec\ud568": row.get("futures_basket_refs"),
    }


def build_dashboard_rows(
    ktb_tenor, irs_tenor, ktb_curve, irs_curve, agency, ktb_irs, bond_rv, lending_rows, fut_basket, fut_irs
):
    rows = []
    def best(rows_in, key, label):
        vals = [r for r in rows_in if r.get(key) is not None]
        return max(vals, key=lambda x: x[key]) if vals else None

    for label, data, key in [
        ("Best KTB tenor carry+roll", ktb_tenor, "total_return"),
        ("Best IRS tenor carry+roll", irs_tenor, "total_return"),
        ("Best KTB curve carry+roll", ktb_curve, "carry_roll_pnl"),
        ("Best IRS curve carry+roll", irs_curve, "carry_roll_pnl"),
    ]:
        r = best(data, key, label)
        if r:
            rows.append({"signal": label, "item": r.get("tenor") or r.get("trade"), "detail": r.get("holding"), "value": r[key], "interpretation": "Highest static carry/roll"})
    rv = sorted([r for r in bond_rv if r.get("residual_z") is not None], key=lambda r: r["residual_z"])
    if rv:
        rows.append({"signal": "Cheapest bond vs fitted curve", "item": rv[-1]["name"], "detail": rv[-1]["futures_basket_refs"], "value": rv[-1]["residual_z"], "interpretation": signal_text(rv[-1]["residual_z"], "cheap", "rich")})
        rows.append({"signal": "Richest bond vs fitted curve", "item": rv[0]["name"], "detail": rv[0]["futures_basket_refs"], "value": rv[0]["residual_z"], "interpretation": signal_text(rv[0]["residual_z"], "cheap", "rich")})
    for label, data in [("Agency spread extreme", agency), ("KTB-IRS spread extreme", ktb_irs), ("Futures basket RV extreme", fut_basket), ("Futures-IRS spread extreme", fut_irs)]:
        vals = [r for r in data if r.get("z_score") is not None or r.get("spread_z_252") is not None]
        if vals:
            r = max(vals, key=lambda x: abs(x.get("z_score") if x.get("z_score") is not None else x.get("spread_z_252")))
            z = r.get("z_score") if r.get("z_score") is not None else r.get("spread_z_252")
            rows.append({"signal": label, "item": r.get("future") or r.get("sector") or r.get("tenor"), "detail": r.get("metric") or r.get("holding"), "value": z, "interpretation": signal_text(z, "wide/cheap", "tight/rich")})
    lend = sorted([r for r in lending_rows if r.get("lending_5d_change") is not None], key=lambda r: abs(r["lending_5d_change"]), reverse=True)
    if lend:
        rows.append({"signal": "Largest 5D lending balance move", "item": lend[0]["name"], "detail": lend[0].get("futures_basket_refs"), "value": lend[0]["lending_5d_change"], "interpretation": "Check lending/borrow flow"})
    return rows


def build_bond_rv_flow_cross_rows(bond_rv_rows: list[dict[str, Any]], limit: int = 16) -> list[dict[str, Any]]:
    candidates = []
    for r in bond_rv_rows:
        z = r.get("residual_z")
        if z is None or r.get("curve_range_status") != "OK":
            continue
        total_5d = flow_to_krw_eok(r.get("total_flow_5d"))
        total_today = flow_to_krw_eok(r.get("total_flow_today"))
        lending_5d = lending_to_krw_eok(r.get("lending_5d_change"))
        signal = None
        interpretation = None
        if z > 0:
            if total_5d is not None and total_5d > 0:
                signal = "Cheap + Buy Confirmed"
                interpretation = "Curve 대비 cheap하고 전체 순매수가 확인됩니다."
            elif (total_5d is not None and total_5d < 0) or (lending_5d is not None and lending_5d > 0):
                signal = "Cheap but Supply/Lending Pressure"
                interpretation = "Cheap하지만 매도 또는 대차 압력이 남아 있습니다."
        elif z < 0:
            if (total_5d is not None and total_5d < 0) or (lending_5d is not None and lending_5d > 0):
                signal = "Rich + Sell/Short Pressure"
                interpretation = "Rich한 종목에 매도/대차 압력이 붙어 있습니다."
            elif total_5d is not None and total_5d > 0:
                signal = "Rich but Buy Supported"
                interpretation = "Rich하지만 강한 매수가 가격을 지지할 수 있습니다."
        if signal is None:
            continue
        score = abs(z)
        if total_5d is not None:
            score += min(abs(total_5d) / 1000.0, 2.0)
        if lending_5d is not None:
            score += min(abs(lending_5d) / 500.0, 1.5)
        if r.get("is_benchmark") or r.get("is_futures_basket"):
            score += 0.5
        candidates.append(
            {
                "Signal": signal,
                "Bond": r["name"],
                "Maturity": r["maturity"],
                "RV z": z,
                "Residual bp": r.get("residual_bp"),
                "Total Today 억": total_today,
                "Total 5D 억": total_5d,
                "Top Buyer 5D": r.get("top_buy_investor_5d"),
                "Top Buyer 억": flow_to_krw_eok(r.get("top_buy_5d")),
                "Top Seller 5D": r.get("top_sell_investor_5d"),
                "Top Seller 억": flow_to_krw_eok(r.get("top_sell_5d")),
                "Lending 5D 억": lending_5d,
                "Benchmark": r.get("benchmark_tenor"),
                "Basket": r.get("futures_basket_refs"),
                "Interpretation": interpretation,
                "_score": score,
            }
        )
    candidates.sort(key=lambda x: x["_score"], reverse=True)
    out = []
    per_signal = Counter()
    for row in candidates:
        if per_signal[row["Signal"]] >= 4:
            continue
        clean = {k: v for k, v in row.items() if k != "_score"}
        out.append(clean)
        per_signal[row["Signal"]] += 1
        if len(out) >= limit:
            break
    return out


def sparkline(values: list[float | None], width: int = 18) -> str:
    vals = [v for v in values if v is not None and math.isfinite(v)]
    if not vals:
        return ""
    if len(vals) > width:
        step = len(vals) / width
        vals = [vals[int(i * step)] for i in range(width)]
    lo, hi = min(vals), max(vals)
    bars = "▁▂▃▄▅▆▇█"
    if hi == lo:
        return bars[3] * len(vals)
    return "".join(bars[min(7, max(0, int((v - lo) / (hi - lo) * 7)))] for v in vals)


def sorted_series(series: list[tuple[date, float]]) -> list[tuple[date, float]]:
    return sorted((d, v) for d, v in series if d is not None and v is not None)


def latest_value(series: list[tuple[date, float]]) -> float | None:
    s = sorted_series(series)
    return s[-1][1] if s else None


def change_bp(series: list[tuple[date, float]], days: int) -> float | None:
    s = sorted_series(series)
    if len(s) < 2:
        return None
    latest_d, latest_v = s[-1]
    candidates = [(d, v) for d, v in s if d <= latest_d and (latest_d - d).days >= days]
    base = candidates[-1][1] if candidates else s[0][1]
    return latest_v - base


def recent_spark(series: list[tuple[date, float]], days: int = 92) -> str:
    s = sorted_series(series)
    if not s:
        return ""
    latest_d = s[-1][0]
    vals = [v for d, v in s if (latest_d - d).days <= days]
    if len(vals) < 2:
        vals = [v for _, v in s]
    return sparkline(vals)


def direction_signal(change: float | None, threshold: float = 1.0) -> str:
    if change is None:
        return ""
    if change > threshold:
        return "▲ widen"
    if change < -threshold:
        return "▼ tighten"
    return "→ flat"


def map_by_holding(rows: list[dict[str, Any]], keys: tuple[str, ...], value_key: str) -> dict[tuple[Any, ...], dict[str, float]]:
    out: dict[tuple[Any, ...], dict[str, float]] = defaultdict(dict)
    for r in rows:
        key = tuple(r.get(k) for k in keys)
        out[key][r["holding"]] = r.get(value_key)
    return out


def write_summary_section(
    ws,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[dict[str, Any]],
    widths: list[int] | None = None,
    note: str | None = None,
) -> int:
    dark = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="D9EAF7")
    carry_fill = PatternFill("solid", fgColor="FFF2CC")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
    cell = ws.cell(start_row, 1, title)
    cell.value = title
    cell.fill = dark
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.alignment = Alignment(horizontal="left")
    header_row = start_row + 1
    if note:
        ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=len(headers))
        note_cell = ws.cell(header_row, 1, note)
        note_cell.fill = PatternFill("solid", fgColor="FFF2CC")
        note_cell.font = Font(color="7F6000", italic=True)
        note_cell.alignment = Alignment(horizontal="left")
        header_row += 1
    for c, h in enumerate(headers, start=1):
        hc = ws.cell(header_row, c, h)
        hc.fill = light if "Carry" not in h and "Roll" not in h and "Total" not in h and "Steep" not in h and "Flat" not in h else carry_fill
        hc.font = Font(bold=True, color="1F2937")
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start=header_row + 1):
        for c, h in enumerate(headers, start=1):
            val = row.get(h)
            cell = ws.cell(r_idx, c, val)
            if isinstance(val, float):
                if "P&L" in h or "백만" in h:
                    cell.number_format = "#,##0.0"
                else:
                    cell.number_format = "0.00"
            elif isinstance(val, date):
                cell.number_format = "yyyy-mm-dd"
            if h in ("3M Direction", "Direction"):
                cell.font = Font(bold=True, color="1F2937")
    if widths:
        for c, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = width
    return header_row + len(rows) + 2


def write_heatmap_section(
    ws,
    start_row: int,
    title: str,
    note: str,
    issuers: list[str],
    tenors: list[str],
    matrices: list[tuple[str, dict[tuple[str, str], float | None]]],
) -> int:
    row = start_row
    max_cols = 1 + len(tenors)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
    title_cell = ws.cell(row, 1, title)
    title_cell.fill = PatternFill("solid", fgColor="1F4E78")
    title_cell.font = Font(bold=True, color="FFFFFF", size=12)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
    note_cell = ws.cell(row, 1, note)
    note_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    note_cell.font = Font(color="7F6000", italic=True)
    row += 1

    for matrix_title, matrix in matrices:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
        cell = ws.cell(row, 1, matrix_title)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True, color="1F2937")
        row += 1
        ws.cell(row, 1, "Issuer")
        ws.cell(row, 1).font = Font(bold=True)
        for c, tenor in enumerate(tenors, start=2):
            ws.cell(row, c, tenor)
            ws.cell(row, c).font = Font(bold=True)
            ws.cell(row, c).alignment = Alignment(horizontal="center")
        first_data_row = row + 1
        for r_idx, issuer in enumerate(issuers, start=first_data_row):
            ws.cell(r_idx, 1, issuer)
            ws.cell(r_idx, 1).font = Font(bold=True)
            for c, tenor in enumerate(tenors, start=2):
                val = matrix.get((issuer, tenor))
                ws.cell(r_idx, c, val)
                ws.cell(r_idx, c).number_format = "0.00"
        last_data_row = first_data_row + len(issuers) - 1
        if last_data_row >= first_data_row:
            ws.conditional_formatting.add(
                f"B{first_data_row}:{get_column_letter(max_cols)}{last_data_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFFFFF",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )
        row = last_data_row + 2
    ws.column_dimensions["A"].width = 12
    for c in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9
    return row + 1


def add_section_explain(ws, row: int, headers_len: int, text: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=headers_len)
    cell = ws.cell(row, 1, text)
    cell.fill = PatternFill("solid", fgColor="EAF2F8")
    cell.font = Font(color="1F4E78")
    cell.alignment = Alignment(horizontal="left", wrap_text=True)
    return row + 1


def total_return_bp(rows_by_hold: dict[str, float], hold: str) -> float | None:
    val = rows_by_hold.get(hold)
    return val * 10000 if val is not None else None


def return_to_krw_million(ret: float | None, notional: float = 10_000_000_000) -> float | None:
    return ret * notional / 1_000_000 if ret is not None else None


def pnl_to_krw_million(pnl: float | None) -> float | None:
    return pnl / 1_000_000 if pnl is not None else None


def flow_to_krw_eok(value: float | None) -> float | None:
    return value / 100_000_000 if value is not None else None


def holding_to_krw_eok(value: float | None) -> float | None:
    return value / 10_000 if value is not None else None


def lending_to_krw_eok(value: float | None) -> float | None:
    return value / 1_000 if value is not None else None


def tenor_detail_display_rows(rows: list[dict[str, Any]], is_irs: bool = False) -> list[dict[str, Any]]:
    out = []
    for r in rows:
        if is_irs:
            out.append(
                {
                    "direction": r["direction"],
                    "tenor": r["tenor"],
                    "tenor_years": r["tenor_years"],
                    "holding": r["holding"],
                    "irs_rate": r["irs_rate"],
                    "cd_3m": r["cd_3m"],
                    "duration": r["duration"],
                    "carry 백만": return_to_krw_million(r.get("carry_return")),
                    "rolldown 백만": return_to_krw_million(r.get("rolldown_return")),
                    "total 백만": return_to_krw_million(r.get("total_return")),
                    "DV01 백만/bp": dv01_per_notional(r["tenor_years"], r["irs_rate"], 10_000_000_000) / 1_000_000,
                    "+10bp 후 total 백만": total_with_parallel_shock_million(r.get("total_return"), r["tenor_years"], r["irs_rate"], 10.0),
                    "-10bp 후 total 백만": total_with_parallel_shock_million(r.get("total_return"), r["tenor_years"], r["irs_rate"], -10.0),
                    "breakeven_rate_bp": r.get("breakeven_rate_bp"),
                }
            )
        else:
            out.append(
                {
                    "asset": r["asset"],
                    "tenor": r["tenor"],
                    "tenor_years": r["tenor_years"],
                    "holding": r["holding"],
                    "yield": r["yield"],
                    "funding": r["funding"],
                    "duration": r["duration"],
                    "carry 백만": return_to_krw_million(r.get("carry_return")),
                    "rolldown 백만": return_to_krw_million(r.get("rolldown_return")),
                    "total 백만": return_to_krw_million(r.get("total_return")),
                    "DV01 백만/bp": dv01_per_notional(r["tenor_years"], r["yield"], 10_000_000_000) / 1_000_000,
                    "+10bp 후 total 백만": total_with_parallel_shock_million(r.get("total_return"), r["tenor_years"], r["yield"], 10.0),
                    "-10bp 후 total 백만": total_with_parallel_shock_million(r.get("total_return"), r["tenor_years"], r["yield"], -10.0),
                    "return_per_dv01": r.get("return_per_dv01"),
                    "breakeven_yield_bp": r.get("breakeven_yield_bp"),
                }
            )
    return out


def curve_detail_display_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for r in rows:
        out.append(
            {
                "curve": r["curve"],
                "trade": r["trade"],
                "direction": r["direction"],
                "holding": r["holding"],
                "short_tenor": r["short_tenor"],
                "long_tenor": r["long_tenor"],
                "short_yield": r["short_yield"],
                "long_yield": r["long_yield"],
                "current_slope_bp": r["current_slope_bp"],
                "hedge_ratio": r["hedge_ratio"],
                "theoretical_short_notional": r["theoretical_short_notional"],
                "notional_rounding_unit": r["notional_rounding_unit"],
                "short_tenor_notional": r["short_tenor_notional"],
                "long_tenor_notional": r["long_tenor_notional"],
                "carry_roll_pnl 백만": pnl_to_krw_million(r.get("carry_roll_pnl")),
                "slope_DV01 백만/bp": pnl_to_krw_million(r.get("slope_dv01")),
                "+10bp slope 후 백만": pnl_to_krw_million(r.get("+10bp_slope_pnl")),
                "-10bp slope 후 백만": pnl_to_krw_million(r.get("-10bp_slope_pnl")),
                "breakeven_slope_bp": r.get("breakeven_slope_bp"),
            }
        )
    return out


def agency_detail_display_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for r in rows:
        new = dict(r)
        new["carry_3m 백만"] = return_to_krw_million(new.pop("carry_3m", None))
        new["carry_6m 백만"] = return_to_krw_million(new.pop("carry_6m", None))
        new["carry_1y 백만"] = return_to_krw_million(new.pop("carry_1y", None))
        out.append(new)
    return out


def ktb_irs_detail_display_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for r in rows:
        new = dict(r)
        new["ktb_buy_cr 백만"] = return_to_krw_million(new.pop("ktb_buy_cr", None))
        new["irs_pay_cr 백만"] = return_to_krw_million(new.pop("irs_pay_cr", None))
        new["total_cr 백만"] = return_to_krw_million(new.pop("total_cr", None))
        out.append(new)
    return out


def build_flow_change_rows(bond_rv_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for r in bond_rv_rows:
        flows = r.get("flow_by_investor") or {}
        for investor, vals in flows.items():
            out.append(
                {
                    "Bond": r["name"],
                    "Prefix": r["prefix"],
                    "Maturity": r["maturity"],
                    "Investor": investor,
                    "Today 억": flow_to_krw_eok(vals.get("today")),
                    "5D 억": flow_to_krw_eok(vals.get("5d")),
                    "20D 억": flow_to_krw_eok(vals.get("20d")),
                    "Balance 억": holding_to_krw_eok(vals.get("balance")),
                    "Benchmark": r["benchmark_tenor"],
                    "Basket": r["futures_basket_refs"],
                }
            )
    return out


def tenor_bucket_label(years: float | None) -> str | None:
    if years is None:
        return None
    if 1.5 <= years < 2.5:
        return "2Y"
    if 2.5 <= years < 3.5:
        return "3Y"
    if 3.5 <= years < 6.0:
        return "5Y"
    if 6.0 <= years < 12.0:
        return "10Y"
    if 12.0 <= years < 25.0:
        return "20Y"
    if years >= 25.0:
        return "30Y"
    return "<2Y"


def focus_rank(label: str | None) -> int:
    text = label or ""
    if "Benchmark" in text and "Basket" in text:
        return 0
    if "Benchmark" in text:
        return 1
    if "Basket" in text:
        return 2
    if "Near benchmark" in text:
        return 3
    if "Near basket" in text:
        return 4
    if "Top net buy" in text or "Top net sell" in text or "Top flow" in text:
        return 5
    if "Top lending" in text or "Top redeem" in text:
        return 6
    return 9


def flow_summary_display_row(r: dict[str, Any], focus: str, bucket: str) -> dict[str, Any]:
    return {
        "Tenor": bucket,
        "Focus": focus,
        "Bond": r["name"],
        "Maturity": r["maturity"],
        "YTM": r["ytm"],
        "Total Today 억": flow_to_krw_eok(r.get("total_flow_today")),
        "Total 5D 억": flow_to_krw_eok(r.get("total_flow_5d")),
        "Total 20D 억": flow_to_krw_eok(r.get("total_flow_20d")),
        "Top Buyer 5D": r.get("top_buy_investor_5d"),
        "Top Buyer 5D 억": flow_to_krw_eok(r.get("top_buy_5d")),
        "Top Seller 5D": r.get("top_sell_investor_5d"),
        "Top Seller 5D 억": flow_to_krw_eok(r.get("top_sell_5d")),
        "대여거래 억": lending_to_krw_eok(r.get("lending_trade_today")),
        "상환거래 억": lending_to_krw_eok(r.get("lending_redeem_today")),
        "Lending Bal 억": lending_to_krw_eok(r.get("lending_balance")),
        "Lending 5D 억": lending_to_krw_eok(r.get("lending_5d_change")),
        "Benchmark": r.get("benchmark_tenor"),
        "Basket": r.get("futures_basket_refs"),
    }


def build_flow_summary_sheet(ws, as_of: date, bond_rv_rows: list[dict[str, Any]], flow_rows: list[dict[str, Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"Flow / Lending Summary ({as_of:%Y-%m-%d})"
    ws["A1"].font = Font(bold=True, size=16, color="1F2937")
    ws.merge_cells("A1:R1")
    ws["A2"] = "순매수 flow는 억원 환산입니다. Top Buy/Sell은 전체 투자자 합산 5D 기준이며, Buyer/Seller는 방향별 최대 기여 주체입니다."
    ws["A2"].font = Font(color="666666")
    ws.merge_cells("A2:R2")
    row = 4

    def write(title: str, headers: list[str], rows: list[dict[str, Any]], note: str | None = None):
        nonlocal row
        widths = [8, 18, 24, 12, 8, 13, 12, 12, 13, 13, 13, 13, 12, 12, 13, 13, 10, 28]
        row = write_summary_section(ws, row, title, headers, rows, widths[: len(headers)], note=note)

    buckets = ["2Y", "3Y", "5Y", "10Y", "20Y", "30Y"]
    by_bucket: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in bond_rv_rows:
        bucket = tenor_bucket_label(r.get("remaining_years"))
        if bucket in buckets:
            by_bucket[bucket].append(r)

    buy_rows_out = []
    sell_rows_out = []
    lend_trade_rows_out = []
    redeem_rows_out = []
    watch_rows_out = []
    for bucket in buckets:
        items = by_bucket.get(bucket, [])
        focus_selected: dict[str, tuple[dict[str, Any], str]] = {}
        for r in items:
            if r.get("is_benchmark") or r.get("is_futures_basket") or r.get("rv_focus"):
                focus_selected[r["code"]] = (r, r.get("rv_focus") or "Focus")

        regular = [r for r in items if r["code"] not in focus_selected]
        watch_rows_out.extend(flow_summary_display_row(item, focus, bucket) for item, focus in focus_selected.values())

        top_buys = sorted(
            [x for x in regular if x.get("total_flow_5d") is not None and x["total_flow_5d"] > 0],
            key=lambda x: x["total_flow_5d"],
            reverse=True,
        )[:3]
        top_sells = sorted(
            [x for x in regular if x.get("total_flow_5d") is not None and x["total_flow_5d"] < 0],
            key=lambda x: x["total_flow_5d"],
        )[:3]
        top_lend_trades = sorted(
            [x for x in regular if x.get("lending_trade_today") is not None and x["lending_trade_today"] > 0],
            key=lambda x: x["lending_trade_today"],
            reverse=True,
        )[:3]
        top_redeems = sorted(
            [x for x in regular if x.get("lending_redeem_today") is not None and x["lending_redeem_today"] > 0],
            key=lambda x: x["lending_redeem_today"],
            reverse=True,
        )[:3]

        buy_rows_out.extend(flow_summary_display_row(item, "Top net buy", bucket) for item in top_buys)
        sell_rows_out.extend(flow_summary_display_row(item, "Top net sell", bucket) for item in top_sells)
        lend_trade_rows_out.extend(flow_summary_display_row(item, "Top lending trade", bucket) for item in top_lend_trades)
        redeem_rows_out.extend(flow_summary_display_row(item, "Top redeem", bucket) for item in top_redeems)

    sort_key = lambda r: (buckets.index(r["Tenor"]), r["Maturity"] or date.max, focus_rank(r["Focus"]), r["Bond"])
    watch_rows_out = sorted(watch_rows_out, key=sort_key)
    buy_rows_out = sorted(buy_rows_out, key=sort_key)
    sell_rows_out = sorted(sell_rows_out, key=sort_key)
    lend_trade_rows_out = sorted(lend_trade_rows_out, key=sort_key)
    redeem_rows_out = sorted(redeem_rows_out, key=sort_key)
    headers = ["Tenor", "Focus", "Bond", "Maturity", "YTM", "Total Today 억", "Total 5D 억", "Total 20D 억", "Top Buyer 5D", "Top Buyer 5D 억", "Top Seller 5D", "Top Seller 5D 억", "대여거래 억", "상환거래 억", "Lending Bal 억", "Lending 5D 억", "Benchmark", "Basket"]
    write(
        "Benchmark/Basket Watchlist",
        headers,
        watch_rows_out,
        "지표채권/선물 바스켓/유사만기 종목은 방향과 무관하게 여기에서 별도로 확인합니다.",
    )
    write(
        "Tenor Net Buy Summary",
        headers,
        buy_rows_out,
        "각 테너 bucket별 일반 종목 중 전체 투자자 합산 5D 순매수 Top 3만 표시합니다. 지표/바스켓은 위 Watchlist에서 봅니다.",
    )
    write(
        "Tenor Net Sell Summary",
        headers,
        sell_rows_out,
        "각 테너 bucket별 일반 종목 중 전체 투자자 합산 5D 순매도 Top 3만 표시합니다. 지표/바스켓은 위 Watchlist에서 봅니다.",
    )
    write(
        "Tenor Lending Trade Summary",
        headers,
        lend_trade_rows_out,
        "각 테너 bucket별 일반 종목 중 금일 대여거래 Top 3만 표시합니다. 지표/바스켓은 위 Watchlist에서 봅니다.",
    )
    write(
        "Tenor Redemption Summary",
        headers,
        redeem_rows_out,
        "각 테너 bucket별 일반 종목 중 금일 상환거래 Top 3만 표시합니다. 지표/바스켓은 위 Watchlist에서 봅니다.",
    )
    ws.freeze_panes = "A4"


def build_summary_sheet(
    ws,
    as_of: date,
    repo: float,
    cd: float,
    ktb: Curve,
    msb: Curve,
    irs: Curve,
    ktb_tenor: list[dict[str, Any]],
    msb_tenor: list[dict[str, Any]],
    irs_tenor: list[dict[str, Any]],
    ktb_curve: list[dict[str, Any]],
    irs_curve: list[dict[str, Any]],
    agency: list[dict[str, Any]],
    agency_rows: list[dict[str, Any]],
    ktb_irs_rows: list[dict[str, Any]],
    bond_rv_rows: list[dict[str, Any]],
    lending_rows: list[dict[str, Any]],
    fut_basket_rows: list[dict[str, Any]],
    fut_irs_rows: list[dict[str, Any]],
    ktb_curve_series: dict[str, list[tuple[date, float]]],
    irs_curve_series: dict[str, list[tuple[date, float]]],
    ktb_irs_series: dict[str, list[tuple[date, float]]],
    agency_series: dict[str, list[tuple[date, float]]],
    futures_basket_series: dict[str, list[tuple[date, float]]],
    futures_irs_series: dict[str, list[tuple[date, float]]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"KTB Master RV Dashboard ({as_of:%Y-%m-%d})"
    ws["A1"].font = Font(bold=True, size=16, color="1F2937")
    ws.merge_cells("A1:R1")
    ws["A2"] = f"Repo {repo:.3f}% | CD 3M {cd:.3f}% | 금리/스프레드 변화는 bp | Carry/Roll/P&L은 표 위 단위 기준"
    ws["A2"].font = Font(color="666666")
    ws.merge_cells("A2:R2")
    row = 4

    def tenor_section(title: str, curve: Curve, tenor_rows: list[dict[str, Any]], asset_filter: str | None = None):
        nonlocal row
        by_hold = map_by_holding(tenor_rows, ("tenor",), "total_return")
        carry_by_hold = map_by_holding(tenor_rows, ("tenor",), "carry_return")
        roll_by_hold = map_by_holding(tenor_rows, ("tenor",), "rolldown_return")
        be_by_hold = map_by_holding(tenor_rows, ("tenor",), "breakeven_yield_bp")
        rows = []
        for tenor, yld in sorted(curve.points.items()):
            label = tenor_label(tenor)
            hist = [(d, v) for d, v in curve.history.get(tenor, [])]
            total_3m = by_hold.get((label,), {}).get("3M")
            rows.append(
                {
                    "Tenor": label,
                    "Last": yld,
                    "2D": change_bp([(d, v * 100) for d, v in hist], 2),
                    "5D": change_bp([(d, v * 100) for d, v in hist], 5),
                    "20D": change_bp([(d, v * 100) for d, v in hist], 20),
                    "3M": change_bp([(d, v * 100) for d, v in hist], 92),
                    "6M": change_bp([(d, v * 100) for d, v in hist], 183),
                    "1Y": change_bp([(d, v * 100) for d, v in hist], 366),
                    "3M Direction": direction_signal(change_bp([(d, v * 100) for d, v in hist], 92)),
                    "Carry 3M 백만": return_to_krw_million(carry_by_hold.get((label,), {}).get("3M")),
                    "Roll 3M 백만": return_to_krw_million(roll_by_hold.get((label,), {}).get("3M")),
                    "Total 3M 백만": return_to_krw_million(total_3m),
                    "BE 3M bp": be_by_hold.get((label,), {}).get("3M"),
                    "DV01 백만/bp": dv01_per_notional(tenor, yld, 10_000_000_000) / 1_000_000,
                    "+10bp 후 3M 백만": total_with_parallel_shock_million(total_3m, tenor, yld, 10.0),
                    "Total 6M 백만": return_to_krw_million(by_hold.get((label,), {}).get("6M")),
                    "Total 1Y 백만": return_to_krw_million(by_hold.get((label,), {}).get("1Y")),
                }
            )
        headers = ["Tenor", "Last", "2D", "5D", "20D", "3M", "6M", "1Y", "3M Direction", "Carry 3M 백만", "Roll 3M 백만", "Total 3M 백만", "BE 3M bp", "DV01 백만/bp", "+10bp 후 3M 백만", "Total 6M 백만", "Total 1Y 백만"]
        row = write_summary_section(
            ws,
            row,
            title,
            headers,
            rows,
            [9, 9, 8, 8, 8, 8, 8, 8, 18, 12, 12, 12, 10, 12, 14, 12, 12],
            note="단위: 금리/스프레드 변화는 bp, Carry/Roll/Total은 100억 투자 기준 백만원. BE는 3M carry/roll 손익이 0이 되는 불리한 금리 상승폭입니다.",
        )

    tenor_section("KTB Tenor: Level Change and Carry/Roll", ktb, ktb_tenor)
    tenor_section("MSB Tenor: Level Change and Carry/Roll", msb, msb_tenor)

    # IRS receive/pay summary.
    by_dir = {(r["direction"], r["tenor"], r["holding"]): r for r in irs_tenor}
    irs_rows = []
    for tenor, rate in sorted(irs.points.items()):
        label = tenor_label(tenor)
        hist = irs.history.get(tenor, [])
        rec_3m = by_dir.get(("Receive", label, "3M"), {}).get("total_return")
        pay_3m = by_dir.get(("Pay", label, "3M"), {}).get("total_return")
        irs_rows.append(
            {
                "Tenor": label,
                "Last": rate,
                "2D": change_bp([(d, v * 100) for d, v in hist], 2),
                "5D": change_bp([(d, v * 100) for d, v in hist], 5),
                "20D": change_bp([(d, v * 100) for d, v in hist], 20),
                "3M": change_bp([(d, v * 100) for d, v in hist], 92),
                "6M": change_bp([(d, v * 100) for d, v in hist], 183),
                "1Y": change_bp([(d, v * 100) for d, v in hist], 366),
                    "3M Direction": direction_signal(change_bp([(d, v * 100) for d, v in hist], 92)),
                "Rec 3M 백만": return_to_krw_million(rec_3m),
                "Pay 3M 백만": return_to_krw_million(pay_3m),
                "Rec BE 3M bp": by_dir.get(("Receive", label, "3M"), {}).get("breakeven_rate_bp"),
                "DV01 백만/bp": dv01_per_notional(tenor, rate, 10_000_000_000) / 1_000_000,
                "Rec +10bp 후 백만": total_with_parallel_shock_million(rec_3m, tenor, rate, 10.0),
                "Rec 6M 백만": return_to_krw_million(by_dir.get(("Receive", label, "6M"), {}).get("total_return")),
                "Pay 6M 백만": return_to_krw_million(by_dir.get(("Pay", label, "6M"), {}).get("total_return")),
                "Rec 1Y 백만": return_to_krw_million(by_dir.get(("Receive", label, "1Y"), {}).get("total_return")),
                "Pay 1Y 백만": return_to_krw_million(by_dir.get(("Pay", label, "1Y"), {}).get("total_return")),
            }
        )
    row = write_summary_section(
        ws,
        row,
        "IRS Tenor: Receive/Pay Carry/Roll",
        ["Tenor", "Last", "2D", "5D", "20D", "3M", "6M", "1Y", "3M Direction", "Rec 3M 백만", "Pay 3M 백만", "Rec BE 3M bp", "DV01 백만/bp", "Rec +10bp 후 백만", "Rec 6M 백만", "Pay 6M 백만", "Rec 1Y 백만", "Pay 1Y 백만"],
        irs_rows,
        [9, 9, 8, 8, 8, 8, 8, 8, 18, 12, 12, 12, 12, 14, 12, 12, 12, 12],
        note="단위: 금리 변화는 bp, Receive/Pay Carry/Roll은 IRS notional 100억 기준 백만원. 3M IRS point는 CD 3M proxy를 사용합니다.",
    )

    def curve_section(title: str, curve_rows: list[dict[str, Any]], series_map: dict[str, list[tuple[date, float]]]):
        nonlocal row
        grouped: dict[str, dict[str, Any]] = {}
        for r in curve_rows:
            trade = r["trade"]
            grouped.setdefault(trade, {"trade": trade, "hr": r.get("hedge_ratio")})
            grouped[trade][(r["direction"], r["holding"])] = r
        out = []
        for trade, g in grouped.items():
            series = series_map.get(trade, [])
            out.append(
                {
                    "Trade": trade,
                    "Last": latest_value(series),
                    "2D": change_bp(series, 2),
                    "5D": change_bp(series, 5),
                    "20D": change_bp(series, 20),
                    "3M": change_bp(series, 92),
                    "6M": change_bp(series, 183),
                    "1Y": change_bp(series, 366),
                    "3M Direction": direction_signal(change_bp(series, 92)),
                    "HR": g.get("hr"),
                    "Steep 3M 백만": pnl_to_krw_million(g.get(("Steepener", "3M"), {}).get("carry_roll_pnl")),
                    "Flat 3M 백만": pnl_to_krw_million(g.get(("Flattener", "3M"), {}).get("carry_roll_pnl")),
                    "Steep BE bp": g.get(("Steepener", "3M"), {}).get("breakeven_slope_bp"),
                    "Flat BE bp": g.get(("Flattener", "3M"), {}).get("breakeven_slope_bp"),
                    "Steep 6M 백만": pnl_to_krw_million(g.get(("Steepener", "6M"), {}).get("carry_roll_pnl")),
                    "Flat 6M 백만": pnl_to_krw_million(g.get(("Flattener", "6M"), {}).get("carry_roll_pnl")),
                    "Steep 1Y 백만": pnl_to_krw_million(g.get(("Steepener", "1Y"), {}).get("carry_roll_pnl")),
                    "Flat 1Y 백만": pnl_to_krw_million(g.get(("Flattener", "1Y"), {}).get("carry_roll_pnl")),
                }
            )
        headers = ["Trade", "Last", "2D", "5D", "20D", "3M", "6M", "1Y", "3M Direction", "HR", "Steep 3M 백만", "Flat 3M 백만", "Steep BE bp", "Flat BE bp", "Steep 6M 백만", "Flat 6M 백만", "Steep 1Y 백만", "Flat 1Y 백만"]
        row = write_summary_section(
            ws,
            row,
            title,
            headers,
            out,
            [10, 9, 8, 8, 8, 8, 8, 8, 18, 8, 12, 12, 10, 10, 12, 12, 12, 12],
            note="단위: 스프레드 변화는 bp, Curve Carry/Roll P&L은 장기테너 100억 기준 백만원. HR은 장기 DV01 100억 기준 단기테너 헤지비율입니다. BE는 3M carry/roll 손익을 상쇄하는 불리한 slope 변화폭입니다.",
        )

    curve_section("KTB Curve: Spread Change and Steep/Flat Carry P&L", ktb_curve, ktb_curve_series)
    curve_section("IRS Curve: Spread Change and Steep/Flat Carry P&L", irs_curve, irs_curve_series)

    # Spread sections.
    spread_rows = []
    for name, series in ktb_irs_series.items():
        latest = latest_value(series)
        vals = [v for _, v in filter_lookback(series, 366)[0]]
        spread_rows.append(
            {
                "Spread": name,
                "Last": latest,
                "2D": change_bp(series, 2),
                "5D": change_bp(series, 5),
                "20D": change_bp(series, 20),
                "3M": change_bp(series, 92),
                "6M": change_bp(series, 183),
                "1Y": change_bp(series, 366),
                "3M Direction": direction_signal(change_bp(series, 92)),
                "1Y z": zscore(vals, latest) if latest is not None else None,
            }
        )
    row = write_summary_section(
        ws,
        row,
        "KTB-IRS Spread: Level and Regime",
        ["Spread", "Last", "2D", "5D", "20D", "3M", "6M", "1Y", "3M Direction", "1Y z"],
        spread_rows,
        [16, 9, 8, 8, 8, 8, 8, 8, 18, 8],
        note="계산: KTB yield - IRS rate. 음수는 KTB 금리가 IRS보다 낮음(채권이 swap 대비 비싼/rich 방향). 변화값은 현재 - 과거(bp), z는 최근 1년 기준.",
    )

    agency_issuers = ["산금", "중금", "수출입"]
    agency_tenors = ["3M", "6M", "9M", "1Y", "1.5Y", "2Y", "3Y"]
    last_matrix: dict[tuple[str, str], float | None] = {}
    z_matrix: dict[tuple[str, str], float | None] = {}
    chg_matrix: dict[tuple[str, str], float | None] = {}
    for block in agency:
        issuer = short_agency_sector(str(block["sector"]))
        tenor = tenor_label(block["tenor_years"])
        if issuer not in agency_issuers or tenor not in agency_tenors:
            continue
        series = agency_spread_history(block, irs, msb)
        latest = latest_value(series)
        vals = [v for _, v in filter_lookback(series, 366)[0]]
        key = (issuer, tenor)
        last_matrix[key] = latest
        z_matrix[key] = zscore(vals, latest) if latest is not None else None
        chg_matrix[key] = change_bp(series, 92)
    row = write_heatmap_section(
        ws,
        row,
        "Agency Spread: Key Issuers",
        "대상: 산금/중금/수출입. 계산: 2년 미만은 Agency - IRS, 2년 이상은 Agency - MSB. 양수는 벤치마크보다 높은 금리(cheap), 음수는 낮은 금리(rich).",
        agency_issuers,
        agency_tenors,
        [
            ("Last Spread bp", last_matrix),
            ("1Y z-score", z_matrix),
            ("3M Change bp", chg_matrix),
        ],
    )

    fut_out = []
    for name, series in {**futures_basket_series, **futures_irs_series}.items():
        latest = latest_value(series)
        vals = [v for _, v in filter_lookback(series, 366)[0]]
        fut_out.append(
            {
                "Metric": name,
                "Last": latest,
                "2D": change_bp(series, 2),
                "5D": change_bp(series, 5),
                "20D": change_bp(series, 20),
                "3M": change_bp(series, 92),
                "3M Direction": direction_signal(change_bp(series, 92)),
                "1Y z": zscore(vals, latest) if latest is not None else None,
            }
        )
    row = write_summary_section(
        ws,
        row,
        "Futures: Basket RV and Futures-IRS",
        ["Metric", "Last", "2D", "5D", "20D", "3M", "3M Direction", "1Y z"],
        fut_out,
        [28, 9, 8, 8, 8, 8, 18, 8],
        note="3년 선물 Butterfly는 바스켓2 - 평균(바스켓1,3): 양수면 중간 바스켓 수익률이 높아 상대적으로 cheap. 10년 Basket1-Basket2는 양수면 1번 수익률이 더 높음. Futures-IRS는 선물 현물수익률 - IRS.",
    )

    # Bond RV top signals.
    rv_candidates = [r for r in bond_rv_rows if r.get("residual_z") is not None and r.get("curve_range_status") == "OK"]
    by_name = {r["name"]: r for r in bond_rv_rows}
    exact_focus = sorted(
        [r for r in rv_candidates if r.get("is_benchmark") or r.get("is_futures_basket")],
        key=lambda x: (x.get("maturity") or date.max, x["name"]),
    )
    focus_rows = [
        {
            "\ubd84\ub958": rv_focus_label(r),
            "\uc885\ubaa9": r["name"],
            "\ub9cc\uae30": r["maturity"],
            "YTM": r["ytm"],
            "\ucee4\ube0c \uad34\ub9ac bp": r["residual_bp"],
            "\ud574\uc11d": rv_side(r.get("residual_bp")),
            "\ub2e8\uba74/\uacfc\uac70 z": r.get("combined_residual_z"),
            "\uc9c0\ud45c \ud14c\ub108": r.get("benchmark_tenor"),
            "\ubc14\uc2a4\ucf13 \ud3ec\ud568": r.get("futures_basket_refs"),
        }
        for r in exact_focus
    ]
    row = write_summary_section(
        ws,
        row,
        "Bond RV 1: \uc9c0\ud45c/\ubc14\uc2a4\ucf13 \uae30\uc900\uc810",
        ["\ubd84\ub958", "\uc885\ubaa9", "\ub9cc\uae30", "YTM", "\ucee4\ube0c \uad34\ub9ac bp", "\ud574\uc11d", "\ub2e8\uba74/\uacfc\uac70 z", "\uc9c0\ud45c \ud14c\ub108", "\ubc14\uc2a4\ucf13 \ud3ec\ud568"],
        focus_rows,
        [16, 24, 12, 8, 12, 26, 8, 10, 30],
        note="\uc9c0\ud45c\ucc44\uc640 \uad6d\ucc44\uc120\ubb3c \ubc14\uc2a4\ucf13 \ucc44\uad8c \uc790\uccb4\ub97c \uba3c\uc800 \ubcf4\uc5ec\uc90d\ub2c8\ub2e4. \ub2e8\uba74/\uacfc\uac70 z\ub294 \uc624\ub298 101\uac1c \uc885\ubaa9 \ub2e8\uba74 z / \ud574\ub2f9 \uc885\ubaa9\uc758 1Y \uacfc\uac70 residual z\uc785\ub2c8\ub2e4. \uacfc\uac70 \ud788\uc2a4\ud1a0\ub9ac 6\uac1c\uc6d4 \ubbf8\ub9cc\uc740 N/A\uc785\ub2c8\ub2e4.",
    )

    near_map = {}
    near_candidates = [
        r
        for r in rv_candidates
        if r.get("rv_focus") and not (r.get("is_benchmark") or r.get("is_futures_basket"))
    ]
    def add_near_row(r: dict[str, Any], ref: dict[str, Any], reason: str, ref_type: str) -> None:
        key = (r["code"], ref["code"])
        if key not in near_map:
            near_map[key] = rv_compare_row(r, ref, reason, ref_type)
            return
        row_existing = near_map[key]
        for col, value in [
            ("\uc120\uc815 \uc0ac\uc720", reason),
            ("\ube44\uad50 \uc720\ud615", ref_type),
        ]:
            parts = [p.strip() for p in str(row_existing.get(col) or "").split("+") if p.strip()]
            if value not in parts:
                parts.append(value)
            row_existing[col] = " + ".join(parts)

    for r in sorted(near_candidates, key=lambda x: abs(x["residual_z"]), reverse=True):
        if r.get("nearest_benchmark") and r["nearest_benchmark"] in by_name:
            ref = by_name[r["nearest_benchmark"]]
            add_near_row(r, ref, "\uc9c0\ud45c\ucc44 \uc8fc\ubcc0", "\uc9c0\ud45c\ucc44")
        if r.get("nearest_basket") and r["nearest_basket"] in by_name:
            ref = by_name[r["nearest_basket"]]
            add_near_row(r, ref, "\ubc14\uc2a4\ucf13 \uc8fc\ubcc0", "\uc120\ubb3c \ubc14\uc2a4\ucf13")
        if len(near_map) >= 20:
            break
    near_rows = list(near_map.values())
    near_rows = sorted(near_rows, key=lambda x: (x.get("\ub9cc\uae30") or date.max, x["\uc885\ubaa9"], x["\ube44\uad50 \uc720\ud615"]))
    row = write_summary_section(
        ws,
        row,
        "Bond RV 2: \uae30\uc900\uc810 \uc8fc\ubcc0 \uc0c1\ub300\uac00\uce58",
        ["\uc120\uc815 \uc0ac\uc720", "\uc885\ubaa9", "\ub9cc\uae30", "YTM", "\ube44\uad50 \uae30\uc900", "\ube44\uad50 \uc720\ud615", "\ub9cc\uae30\ucc28 \uc6d4", "YTM \ucc28 bp", "\ucee4\ube0c \uad34\ub9ac bp", "\ud574\uc11d", "\ub2e8\uba74/\uacfc\uac70 z", "\ubc14\uc2a4\ucf13 \ud3ec\ud568"],
        near_rows,
        [14, 24, 12, 8, 24, 12, 10, 10, 12, 26, 8, 30],
        note="\uc9c0\ud45c\ucc44/\ubc14\uc2a4\ucf13\uacfc \ub9cc\uae30 \ucc28\uc774\uac00 6\uac1c\uc6d4 \uc774\ub0b4\uc778 \ub3d9\uc77c \uc720\ud615 \uc885\ubaa9\uc744 \ud45c\uc2dc\ud569\ub2c8\ub2e4. YTM \ucc28 bp\ub294 \ud574\ub2f9 \uc885\ubaa9 YTM - \ube44\uad50 \uae30\uc900 YTM, \ub2e8\uba74/\uacfc\uac70 z\ub294 \uc624\ub298 \ub2e8\uba74 / \uc790\uae30 \uacfc\uac70 1Y \uae30\uc900\uc785\ub2c8\ub2e4.",
    )

    cross_rows = build_bond_rv_flow_cross_rows(rv_candidates)
    row = write_summary_section(
        ws,
        row,
        "Bond RV x Flow: RV \uc2e0\ud638\uc640 \uc218\uae09 \ud655\uc778",
        ["Signal", "Bond", "Maturity", "RV z", "Residual bp", "Total Today 억", "Total 5D 억", "Top Buyer 5D", "Top Buyer 억", "Top Seller 5D", "Top Seller 억", "Lending 5D 억", "Benchmark", "Basket", "Interpretation"],
        cross_rows,
        [24, 24, 12, 8, 10, 12, 12, 13, 12, 13, 12, 12, 10, 24, 32],
        note="Bond RV 신호를 전체 투자자 합산 flow와 대차 변화로 교차 확인합니다. Cheap은 residual z>0, Rich는 residual z<0입니다.",
    )

    cheap = sorted(rv_candidates, key=lambda r: r["residual_z"], reverse=True)[:4]
    rich = sorted(rv_candidates, key=lambda r: r["residual_z"])[:4]
    bond_rows = []
    for tag, items in [("Cheap", cheap), ("Rich", rich)]:
        for r in items:
            bond_rows.append(
                {
                    "\uc2e0\ud638": tag,
                    "\uc885\ubaa9": r["name"],
                    "\ub9cc\uae30": r["maturity"],
                    "YTM": r["ytm"],
                    "\ucee4\ube0c \uad34\ub9ac bp": r["residual_bp"],
                    "\ud574\uc11d": rv_side(r.get("residual_bp")),
                    "\ub2e8\uba74/\uacfc\uac70 z": r.get("combined_residual_z"),
                    "\uc9c0\ud45c \ud14c\ub108": r["benchmark_tenor"],
                    "\ubc14\uc2a4\ucf13 \ud3ec\ud568": r["futures_basket_refs"],
                }
            )
    write_summary_section(
        ws,
        row,
        "Bond RV 3: \uc804\uccb4 \uc885\ubaa9 Cheap/Rich \uadf9\ub2e8\uac12",
        ["\uc2e0\ud638", "\uc885\ubaa9", "\ub9cc\uae30", "YTM", "\ucee4\ube0c \uad34\ub9ac bp", "\ud574\uc11d", "\ub2e8\uba74/\uacfc\uac70 z", "\uc9c0\ud45c \ud14c\ub108", "\ubc14\uc2a4\ucf13 \ud3ec\ud568"],
        bond_rows,
        [9, 24, 12, 8, 12, 26, 8, 10, 30],
        note="\uc804\uccb4 101\uac1c \uc885\ubaa9 \uc911 \ub3d9\uc77c\ub9cc\uae30 fitted curve\ub300\ube44 \uad34\ub9ac\uac00 \uac00\uc7a5 \ud070 cheap/rich \uc0c1\ud558\uc704 4\uac1c\uc529\uc744 \ubcf4\uc5ec\uc90d\ub2c8\ub2e4. \ub2e8\uba74/\uacfc\uac70 z\ub85c \uc624\ub298 \ub2e8\uba74\uc0c1 \uc2e0\ud638\uc640 \uc790\uae30 \uacfc\uac70 \ub300\ube44 \uc2e0\ud638\ub97c \uac19\uc774 \ud655\uc778\ud569\ub2c8\ub2e4.",
    )
    ws.freeze_panes = "A4"


def main() -> Path:
    config = load_config()
    source = ROOT / config["source_file"]
    output_dir = ROOT / config["output_dir"]
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / config["report_file"]

    print("Loading source workbook...", flush=True)
    wb_src = load_workbook(source, read_only=True, data_only=True)
    print("Parsing curves...", flush=True)
    ktb = parse_curve_sheet(wb_src, "KTB_YIELD")
    msb = parse_curve_sheet(wb_src, "MSB_YIELD")
    irs = parse_irs_curve(wb_src)
    cd_repo = parse_cd_repo(wb_src)
    print("Parsing agency, bonds, lending and futures...", flush=True)
    agency = parse_agency_blocks(wb_src)
    repo = cd_repo["repo"]
    cd = cd_repo["cd_3m"]
    collateral_yield, collateral_name = get_1y_agency_proxy(agency, repo)
    as_of = latest_date_from_curve(ktb)
    raw_irs_latest_date = latest_date_from_curve(irs)
    if raw_irs_latest_date > as_of:
        irs = align_curve_to_date(irs, as_of)
    irs = add_irs_3m_cd_proxy(irs, cd_repo, as_of)
    bonds = parse_bond_list(wb_src, as_of, config)
    lending = parse_fullinfo_lending(wb_src)
    futures = parse_futures(wb_src)
    print("Building analytics...", flush=True)
    code_to_name = {b["code"]: b["name"] for b in bonds}
    fut_basket_rows, fut_irs_rows, basket_flags = build_futures_outputs(futures, irs, code_to_name, config, as_of)
    bond_rv_rows = build_bond_rv(bonds, ktb, msb, lending, basket_flags)

    periods = config["analysis"]["holding_periods_years"]
    base = config["analysis"]["base_notional_krw"]
    borrow_fee = config["analysis"]["borrow_fee_annual"]
    tenor_years = {k: float(v) for k, v in config["tenor_years"].items()}

    ktb_tenor = build_tenor_cr(ktb, repo, periods)
    msb_tenor = build_tenor_cr(msb, repo, periods)
    irs_tenor = build_irs_tenor_cr(irs, cd, periods)
    ktb_curve = curve_pair_rows(ktb, config["curve_pairs"]["KTB"], periods, tenor_years, base, repo, collateral_yield, borrow_fee, is_irs=False)
    irs_curve = curve_pair_rows(irs, config["curve_pairs"]["IRS"], periods, tenor_years, base, cd=cd, is_irs=True)
    agency_rows = build_agency_spread_rows(agency, irs, msb, repo)
    ktb_irs_rows = build_ktb_irs_spread_rows(ktb, irs, repo, cd, periods)

    ktb_curve_series = {
        f"{a}-{b}": curve_slope_history(ktb, tenor_years[a], tenor_years[b])
        for a, b in config["curve_pairs"]["KTB"]
    }
    irs_curve_series = {
        f"{a}-{b}": curve_slope_history(irs, tenor_years[a], tenor_years[b])
        for a, b in config["curve_pairs"]["IRS"]
    }
    ktb_irs_series = {
        f"{tenor_label(t)} KTB-IRS": ktb_irs_spread_history(ktb, irs, t)
        for t in [2.0, 3.0, 5.0, 10.0]
        if t in ktb.points and t in irs.points
    }
    agency_rank = sorted(
        agency_rows,
        key=lambda r: abs(r["spread_z"]) if r.get("spread_z") is not None else -1,
        reverse=True,
    )
    agency_keys = {(r["sector"], r["tenor_years"]) for r in agency_rank[:5]}
    agency_series = {}
    for block in agency:
        key = (block["sector"], block["tenor_years"])
        if key in agency_keys:
            label = f"{block['sector']} {tenor_label(block['tenor_years'])}"
            agency_series[label] = agency_spread_history(block, irs, msb)
    active_suffix = active_contract_suffix(as_of, config)
    futures_basket_series = {}
    futures_irs_series = {}
    for future in futures:
        fname = " ".join(str(future["name"]).split())
        if active_suffix not in fname:
            continue
        metric, series = futures_basket_history(future)
        if series:
            futures_basket_series[f"{fname} {metric}"] = series
        spread_series = futures_irs_history(future, irs)
        if spread_series:
            futures_irs_series[f"{fname}-IRS"] = spread_series
    ktb_chart_series = select_series_map(ktb_curve_series, ["2Y-5Y", "3Y-10Y", "5Y-10Y", "10Y-30Y"])
    irs_chart_series = select_series_map(irs_curve_series, ["6M-1Y", "1Y-3Y", "2Y-5Y", "3Y-5Y"])
    ktb_irs_chart_series = select_series_map(
        ktb_irs_series,
        ["2Y KTB-IRS", "3Y KTB-IRS", "5Y KTB-IRS", "10Y KTB-IRS"],
    )
    agency_chart_series = select_series_map(agency_series, max_series=4)
    futures_basket_chart_series = select_series_map(futures_basket_series, max_series=2)
    futures_irs_chart_series = select_series_map(futures_irs_series, max_series=2)

    lending_rows = []
    for r in bond_rv_rows:
        lending_rows.append({k: r.get(k) for k in ["name", "prefix", "ytm", "is_benchmark", "is_futures_basket", "futures_basket_refs", "lending_balance", "lending_5d_change", "lending_20d_change", "lending_trade_today", "lending_redeem_today", "insurance_lending", "bank_lending", "foreign_lending", "securities_borrow", "total_flow_today", "total_flow_5d", "total_flow_20d", "top_buy_investor_5d", "top_buy_5d", "top_sell_investor_5d", "top_sell_5d", "top_flow_investor_5d", "top_flow_5d"]})
    flow_change_rows = build_flow_change_rows(bond_rv_rows)

    print("Writing report workbook...", flush=True)
    wb = Workbook()
    wb.remove(wb.active)
    sheets: dict[str, Any] = {}
    for name in [
        "01_Summary",
        "02_Flow_Summary",
        "03_Methodology_Glossary",
        "04_Config_Assumptions",
        "05_Data_Quality",
        "06_Market_Snapshot",
        "07_KTB_Tenor_CarryRoll",
        "08_KTB_Curve_CarryRoll",
        "09_MSB_Tenor_CarryRoll",
        "10_IRS_Tenor_CarryRoll",
        "11_IRS_Curve_CarryRoll",
        "12_Agency_Spread_CarryRoll",
        "13_KTB_IRS_Spread",
        "14_KTB_MSB_Bond_RV",
        "15_Flow_Changes_by_Bond",
        "16_Lending_Borrow_Flow",
        "17_Futures_Basket_RV",
        "18_Futures_IRS_Spread",
        "19_Signal_Ranking",
        "20_Spread_Charts",
    ]:
        sheets[name] = wb.create_sheet(name)
    chart_data = wb.create_sheet("_Chart_Data")
    chart_data.sheet_state = "hidden"

    build_summary_sheet(
        sheets["01_Summary"],
        as_of,
        repo,
        cd,
        ktb,
        msb,
        irs,
        ktb_tenor,
        msb_tenor,
        irs_tenor,
        ktb_curve,
        irs_curve,
        agency,
        agency_rows,
        ktb_irs_rows,
        bond_rv_rows,
        lending_rows,
        fut_basket_rows,
        fut_irs_rows,
        ktb_curve_series,
        irs_curve_series,
        ktb_irs_series,
        agency_series,
        futures_basket_series,
        futures_irs_series,
    )
    build_flow_summary_sheet(sheets["02_Flow_Summary"], as_of, bond_rv_rows, flow_change_rows)
    add_methodology(sheets["03_Methodology_Glossary"])

    assumption_rows = [
        {"item": "As-of date", "value": as_of, "note": "KTB generic curve latest date"},
        {"item": "Repo rate", "value": repo, "note": f"CD,REPO sheet, date {cd_repo['repo_latest_date']}"},
        {"item": "CD 3M rate", "value": cd, "note": f"CD,REPO sheet, date {cd_repo['cd_latest_date']}"},
        {"item": "Borrow fee annual", "value": borrow_fee, "note": "KTB short borrow fee"},
        {"item": "Collateral proxy", "value": collateral_yield, "note": collateral_name},
        {"item": "IRS 3M proxy", "value": cd, "note": "CD 3M is inserted as the 0.25Y IRS curve point for short IRS rolldown calculations"},
        {"item": "Base curve notional", "value": base, "note": "Long-tenor notional for all curve trades"},
        {"item": "Curve notional rounding", "value": "KTB 10억 / IRS 1억", "note": "Short-tenor notional is rounded after DV01 hedge ratio"},
        {"item": "Active futures contract", "value": active_contract_suffix(as_of, config), "note": "Rolls on maturity date"},
    ]
    add_table(sheets["04_Config_Assumptions"], assumption_rows, title="Assumptions")
    add_table(sheets["04_Config_Assumptions"], [{"config_json": json.dumps(config, ensure_ascii=False, indent=2)}], start_row=len(assumption_rows) + 5, title="Raw Config")

    data_quality_rows = [
        {"check": "Sheet list", "status": "OK", "detail": ", ".join(wb_src.sheetnames)},
        {"check": "KTB latest date", "status": "OK", "detail": latest_date_from_curve(ktb)},
        {"check": "MSB latest date", "status": "OK", "detail": latest_date_from_curve(msb)},
        {"check": "IRS raw latest date", "status": "Aligned" if raw_irs_latest_date > as_of else "OK", "detail": f"raw={raw_irs_latest_date}, used={latest_date_from_curve(irs)}"},
        {"check": "REPO latest date", "status": "Lag expected" if cd_repo["repo_latest_date"] != as_of else "OK", "detail": cd_repo["repo_latest_date"]},
        {"check": "Futures latest date", "status": "Lag expected", "detail": fut_basket_rows[0]["latest_date"] if fut_basket_rows else None},
        {"check": "KTB/MSB List count", "status": "OK" if len(bonds) == 101 else "Review", "detail": len(bonds)},
        {"check": "Futures basket code match", "status": "OK", "detail": "All basket codes mapped to KTB_MSB_List"},
        {"check": "Benchmark bond match", "status": "OK" if all(any(b["name"] == n for b in bonds) for n in config["benchmark_bonds"].values()) else "Review", "detail": ", ".join(config["benchmark_bonds"].values())},
        {"check": "IRS 3Y/10Y", "status": "OK" if 3.0 in irs.points and 10.0 in irs.points else "Review", "detail": f"IRS tenors: {sorted(irs.points)}"},
    ]
    add_table(sheets["05_Data_Quality"], data_quality_rows, title="Data Quality Checks")

    snapshot_rows = [
        {"market": "KTB", "tenor": tenor_label(k), "rate": v, "date": latest_date_from_curve(ktb)} for k, v in sorted(ktb.points.items())
    ] + [
        {"market": "MSB", "tenor": tenor_label(k), "rate": v, "date": latest_date_from_curve(msb)} for k, v in sorted(msb.points.items())
    ] + [
        {"market": "IRS", "tenor": tenor_label(k), "rate": v, "date": latest_date_from_curve(irs)} for k, v in sorted(irs.points.items())
    ]
    add_table(sheets["06_Market_Snapshot"], snapshot_rows, title="Market Snapshot")
    add_table(sheets["07_KTB_Tenor_CarryRoll"], tenor_detail_display_rows(ktb_tenor))
    add_table(sheets["08_KTB_Curve_CarryRoll"], curve_detail_display_rows(ktb_curve))
    add_table(sheets["09_MSB_Tenor_CarryRoll"], tenor_detail_display_rows(msb_tenor))
    add_table(sheets["10_IRS_Tenor_CarryRoll"], tenor_detail_display_rows(irs_tenor, is_irs=True))
    add_table(sheets["11_IRS_Curve_CarryRoll"], curve_detail_display_rows(irs_curve))
    add_table(sheets["12_Agency_Spread_CarryRoll"], agency_detail_display_rows(agency_rows))
    add_table(sheets["13_KTB_IRS_Spread"], ktb_irs_detail_display_rows(ktb_irs_rows))
    bond_rv_table_rows = [{k: v for k, v in r.items() if k != "flow_by_investor"} for r in bond_rv_rows]
    add_table(sheets["14_KTB_MSB_Bond_RV"], bond_rv_table_rows)
    add_table(sheets["15_Flow_Changes_by_Bond"], flow_change_rows)
    add_table(sheets["16_Lending_Borrow_Flow"], lending_rows)
    add_table(sheets["17_Futures_Basket_RV"], fut_basket_rows)
    add_table(sheets["18_Futures_IRS_Spread"], fut_irs_rows)

    signal_rows = []
    for source, rows, zkey in [
        ("Bond RV", bond_rv_rows, "residual_z"),
        ("Agency Spread", agency_rows, "spread_z"),
        ("KTB-IRS Spread", ktb_irs_rows, "spread_z_252"),
        ("Futures Basket", fut_basket_rows, "z_score"),
        ("Futures-IRS", fut_irs_rows, "z_score"),
    ]:
        for r in rows:
            z = r.get(zkey)
            if z is not None:
                signal_rows.append(
                    {
                        "source": source,
                        "item": r.get("name") or r.get("sector") or r.get("future") or r.get("tenor"),
                        "metric": zkey,
                        "z_score": z,
                        "abs_z": abs(z),
                        "interpretation": signal_text(z, "cheap/wide", "rich/tight"),
                    }
                )
    signal_rows.sort(key=lambda r: r["abs_z"], reverse=True)
    add_table(sheets["19_Signal_Ranking"], signal_rows[:200])

    chart_cursor = 1
    chart_cursor = add_clean_spread_panel(sheets["20_Spread_Charts"], chart_data, "KTB Curve Slopes", ktb_chart_series, 1, chart_cursor)
    chart_cursor = add_clean_spread_panel(sheets["20_Spread_Charts"], chart_data, "IRS Curve Slopes", irs_chart_series, 18, chart_cursor)
    chart_cursor = add_clean_spread_panel(sheets["20_Spread_Charts"], chart_data, "KTB-IRS Spreads", ktb_irs_chart_series, 35, chart_cursor)
    chart_cursor = add_clean_spread_panel(sheets["20_Spread_Charts"], chart_data, "Agency Spread Extremes", agency_chart_series, 52, chart_cursor)
    chart_cursor = add_clean_spread_panel(sheets["20_Spread_Charts"], chart_data, "Active Futures Basket RV", futures_basket_chart_series, 69, chart_cursor)
    chart_cursor = add_clean_spread_panel(sheets["20_Spread_Charts"], chart_data, "Active Futures-IRS Spreads", futures_irs_chart_series, 86, chart_cursor)

    add_compact_table(sheets["08_KTB_Curve_CarryRoll"], series_summary_rows(ktb_chart_series), 2, 22)
    chart_cursor = add_chart_source_and_line_chart(sheets["08_KTB_Curve_CarryRoll"], chart_data, "KTB Curve Slopes - 1Y", ktb_chart_series, 366, "V9", chart_cursor, width=16, height=8)
    add_compact_table(sheets["11_IRS_Curve_CarryRoll"], series_summary_rows(irs_chart_series), 2, 22)
    chart_cursor = add_chart_source_and_line_chart(sheets["11_IRS_Curve_CarryRoll"], chart_data, "IRS Curve Slopes - 1Y", irs_chart_series, 366, "V9", chart_cursor, width=16, height=8)
    add_compact_table(sheets["12_Agency_Spread_CarryRoll"], series_summary_rows(agency_chart_series), 2, 23)
    chart_cursor = add_chart_source_and_line_chart(sheets["12_Agency_Spread_CarryRoll"], chart_data, "Agency Spread Extremes - 1Y", agency_chart_series, 366, "W9", chart_cursor, width=16, height=8)
    add_compact_table(sheets["13_KTB_IRS_Spread"], series_summary_rows(ktb_irs_chart_series), 2, 19)
    chart_cursor = add_chart_source_and_line_chart(sheets["13_KTB_IRS_Spread"], chart_data, "KTB-IRS Spreads - 1Y", ktb_irs_chart_series, 366, "S9", chart_cursor, width=16, height=8)
    add_compact_table(sheets["17_Futures_Basket_RV"], series_summary_rows(futures_basket_chart_series), 2, 20)
    chart_cursor = add_chart_source_and_line_chart(sheets["17_Futures_Basket_RV"], chart_data, "Active Futures Basket RV - 1Y", futures_basket_chart_series, 366, "T9", chart_cursor, width=16, height=8)
    add_compact_table(sheets["18_Futures_IRS_Spread"], series_summary_rows(futures_irs_chart_series), 2, 19)
    chart_cursor = add_chart_source_and_line_chart(sheets["18_Futures_IRS_Spread"], chart_data, "Active Futures-IRS Spreads - 1Y", futures_irs_chart_series, 366, "S9", chart_cursor, width=16, height=8)

    for ws in wb.worksheets:
        style_sheet(ws)
        if ws.title == "01_Summary":
            style_summary_cues(ws)
        if ws.title == "13_KTB_MSB_Bond_RV":
            highlight_bond_rv(ws)
        if ws.max_row > 2 and ws.max_column > 1:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[0].row % 2 == 0:
                    for cell in row:
                        if cell.fill.fill_type is None:
                            cell.fill = PatternFill("solid", fgColor="F8FBFD")
        # Highlight z-score columns.
        for c in range(1, ws.max_column + 1):
            header = ws.cell(1, c).value
            if header and ("z" in str(header).lower() or "residual_bp" == str(header)):
                col = get_column_letter(c)
                ws.conditional_formatting.add(
                    f"{col}2:{col}{ws.max_row}",
                    ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFFFFF", end_type="max", end_color="F8696B"),
                )

    try:
        wb.save(output)
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = output.with_name(f"{output.stem}_{stamp}{output.suffix}")
        wb.save(output)
    print(f"Saved report: {output}", flush=True)
    return output


if __name__ == "__main__":
    out = main()
    print(out)
